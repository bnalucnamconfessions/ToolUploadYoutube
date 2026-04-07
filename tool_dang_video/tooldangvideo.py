# -*- coding: utf-8 -*-
"""
Module đăng video lên YouTube qua YouTube Studio (Selenium).
Cung cấp: init_driver(), upload_video(), upload_videos_batch(), generate_excel()
"""
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, WebDriverException

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    get_column_letter = None


import json

# Thư mục log debug (app.py gọi configure_debug_dir khi khởi động)
_CONFIGURED_DEBUG_DIR = None


def configure_debug_dir(path):
    """Đặt thư mục chứa file debug (vd: .../debug_logs cạnh file exe)."""
    global _CONFIGURED_DEBUG_DIR
    if path and isinstance(path, str) and path.strip():
        _CONFIGURED_DEBUG_DIR = os.path.abspath(path.strip())
        try:
            os.makedirs(_CONFIGURED_DEBUG_DIR, exist_ok=True)
        except OSError:
            pass
    else:
        _CONFIGURED_DEBUG_DIR = None


def _debug_logs_dir():
    """Thư mục ghi log debug — ưu tiên đường dẫn đã cấu hình, sau đó cạnh exe hoặc project."""
    if _CONFIGURED_DEBUG_DIR:
        return _CONFIGURED_DEBUG_DIR
    if getattr(sys, "frozen", False):
        base = os.path.join(os.path.dirname(sys.executable), "debug_logs")
    else:
        base = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir)), "debug_logs")
    try:
        os.makedirs(base, exist_ok=True)
    except OSError:
        pass
    return base


# URL YouTube Studio
YOUTUBE_STUDIO_URL = "https://studio.youtube.com/"
YOUTUBE_UPLOAD_URL = "https://studio.youtube.com/?noapp=1"

# Thư mục profile Chrome mặc định để lưu đăng nhập YouTube (cookie, session)
# (app.py có thể truyền profile_dir khác để chọn nhiều tài khoản)
CHROME_PROFILE_DIR = os.path.join(os.getcwd(), "chrome_youtube_profile")

# Studio đôi khi render multi-progress trong tp-yt-paper-dialog#dialog (class ytcp-multi-progress-monitor)
# thay vì ytcp-uploads-dialog — chỉ query ytcp-uploads-dialog sẽ không thấy #progress-list và chờ mãi.
_JS_YTB_UPLOAD_PROGRESS_ROOT = (
    "function __ytbUploadProgressRoot(){"
    "var pl=document.querySelector('#progress-list');"
    "if(pl){var h=pl.closest('tp-yt-paper-dialog, ytcp-uploads-dialog');if(h)return h;}"
    "var x=document.querySelector('ytcp-uploads-dialog');if(x)return x;"
    "x=document.querySelector('tp-yt-paper-dialog#dialog');if(x)return x;"
    "return document.querySelector('tp-yt-paper-dialog[class*=\"ytcp-multi-progress-monitor\"]');"
    "}"
)


# #region agent log
def _agent_debug_log(hypothesis_id, message, data=None, run_id="init_driver_pre"):
    """
    Ghi log debug dạng NDJSON vào thư mục debug_logs (cạnh exe khi đóng gói PyInstaller).
    Không ghi thông tin nhạy cảm (token, mật khẩu, ...).
    """
    log_path = os.path.join(_debug_logs_dir(), "debug-57c0c7.log")
    payload = {
        "sessionId": "57c0c7",
        "runId": run_id,
        "hypothesisId": hypothesis_id,
        "location": "tooldangvideo.py:init_driver",
        "message": message,
        "data": data or {},
        "timestamp": int(time.time() * 1000),
    }
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        # Tránh làm hỏng luồng chính nếu việc ghi log thất bại
        pass
# #endregion agent log


def _log(log_callback, message):
    if log_callback:
        log_callback(message)


# #region agent log (debug-mode NDJSON)
def _dbg(hypothesis_id: str, message: str, data=None, run_id: str = "speed_profile"):
    """Ghi NDJSON debug (không ghi dữ liệu nhạy cảm)."""
    # Mặc định KHÔNG ghi file debug-*.log (để chạy .exe trên máy khác không sinh log).
    # Bật lại bằng env: YTB_DEBUG_NDJSON=1
    try:
        if str(os.environ.get("YTB_DEBUG_NDJSON", "")).strip() not in ("1", "true", "True", "YES", "yes", "on", "ON"):
            return
    except Exception:
        return
    try:
        payload = {
            "sessionId": "57c0c7",
            "runId": run_id,
            "hypothesisId": hypothesis_id,
            "location": "tool_dang_video/tooldangvideo.py",
            "message": message,
            "data": data or {},
            "timestamp": int(time.time() * 1000),
        }
        try:
            p = os.path.join(_debug_logs_dir(), "debug-speed.ndjson")
            with open(p, "a", encoding="utf-8") as f:
                f.write(json.dumps(payload, ensure_ascii=False) + "\n")
        except Exception:
            pass
    except Exception:
        pass
# #endregion agent log


def _browser_console_debug_verbose():
    """YTB_BROWSER_CONSOLE=1: in thêm log console ở một số bước (ồn ào hơn)."""
    try:
        return str(os.environ.get("YTB_BROWSER_CONSOLE", "")).strip().lower() in (
            "1",
            "true",
            "yes",
            "on",
        )
    except Exception:
        return False


def _is_console_noise(message: str) -> bool:
    """Bỏ qua dòng console thường gặp, không phải lỗi logic upload."""
    m = (message or "").lower()
    noise = (
        "self-xss",
        "self xss",
        "quota_exceeded",
        "gcm",
        "favicon",
        "chrome-extension://",
        "extension://",
        "devtools",
    )
    return any(x in m for x in noise)


def _collect_browser_console_logs(driver, max_entries=120):
    """Đọc buffer log `browser` từ Chrome (cần goog:loggingPrefs trong init_driver)."""
    try:
        raw = driver.get_log("browser")
    except Exception:
        return []
    out = []
    for e in raw[-max_entries:]:
        try:
            msg = (e.get("message") or "")[:2000]
            level = (e.get("level") or "INFO").upper()
            out.append({"level": level, "message": msg, "timestamp": e.get("timestamp")})
        except Exception:
            continue
    return out


def _report_chrome_console_for_debug(driver, log_callback, note=""):
    """
    In ra UI các dòng console SEVERE/WARNING (và một số INFO chứa 'error'), đã lọc nhiễu.
    Gọi khi timeout / lỗi để xem có lỗi JS hay không.
    """
    if not log_callback:
        return
    entries = _collect_browser_console_logs(driver, max_entries=150)
    suffix = f" ({note})" if note else ""
    if not entries:
        _log(
            log_callback,
            f"🔍 Debug Chrome console{suffix}: không đọc được log (hoặc buffer trống).",
        )
        return
    interesting = []
    for e in entries:
        lvl = (e.get("level") or "").upper()
        msg = e.get("message") or ""
        if _is_console_noise(msg):
            continue
        if lvl in ("SEVERE", "WARNING"):
            interesting.append(e)
        elif lvl == "INFO" and "error" in msg.lower():
            interesting.append(e)
    tail = interesting[-18:]
    if not tail:
        _log(
            log_callback,
            f"🔍 Debug Chrome console{suffix}: không thấy SEVERE/WARNING đáng chú ý (đã lọc nhiễu).",
        )
        return
    _log(log_callback, f"🔍 Chrome console — {len(tail)} dòng lỗi/cảnh báo gần đây{suffix}:")
    for e in tail:
        msg = (e.get("message") or "").replace("\n", " ").strip()
        if len(msg) > 500:
            msg = msg[:497] + "..."
        _log(log_callback, f"   [{e.get('level')}] {msg}")
    try:
        if str(os.environ.get("YTB_DEBUG_NDJSON", "")).strip() in (
            "1",
            "true",
            "True",
            "YES",
            "yes",
            "on",
            "ON",
        ):
            _dbg(
                "BR",
                "browser console snapshot",
                {"note": note or "", "n": len(tail), "last": (tail[-1].get("message") or "")[:240]},
            )
    except Exception:
        pass


# Cache nhẹ để tránh set width/header lặp lại (tăng tốc ghi Excel realtime).
# Lưu ý: upload chạy 1 thread, nên cache này an toàn trong phạm vi tool hiện tại.
_EXCEL_WIDTHS_SET = set()   # excel_path
_EXCEL_HEADER_OK = set()    # excel_path


# #region agent log
def _agent_excel_path(excel_filename: str):
    output_dir = os.path.join(os.getcwd(), "output")
    os.makedirs(output_dir, exist_ok=True)
    safe_name = "".join(c for c in str(excel_filename or "") if c.isalnum() or c in "._- ") or "YouTube_Upload_Links.xlsx"
    if not safe_name.endswith(".xlsx"):
        safe_name += ".xlsx"
    return os.path.join(output_dir, safe_name)


def _set_excel_column_widths(ws, num_cols=5):
    """Đặt độ rộng cột để dễ đọc và copy link (STT, Tên file, Link, Thời gian, Trạng thái)."""
    if not OPENPYXL_AVAILABLE or get_column_letter is None:
        return
    widths = [8, 42, 58, 20, 14]  # STT hẹp, Tên file & Link rộng, Thời gian & Trạng thái vừa
    for i, w in enumerate(widths[:num_cols], 1):
        try:
            ws.column_dimensions[get_column_letter(i)].width = w
        except Exception:
            pass


def append_excel_row(file_name: str, url: str, status: str, excel_filename: str = "YouTube_Upload_Links.xlsx", log_callback=None):
    """
    Ghi dần kết quả upload ra Excel ngay khi có kết quả.
    - Nếu file chưa tồn tại: tạo mới + header
    - Nếu đã tồn tại: append thêm 1 dòng
    """
    if not OPENPYXL_AVAILABLE:
        _log(log_callback, "Thiếu thư viện openpyxl, không ghi được Excel theo thời gian thực.")
        return None
    excel_path = _agent_excel_path(excel_filename)
    try:
        # Debug: xác định đang tạo mới hay append
        _agent_debug_log(
            "X1",
            "append_excel_row invoked",
            {"excelPath": excel_path, "exists": os.path.exists(excel_path), "cwd": os.getcwd()},
            run_id="excel_append_pre",
        )
    except Exception:
        pass
    t0 = time.time()
    try:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            # Nếu file cũ chưa có cột Trạng thái thì thêm vào cuối (để tương thích ngược)
            if excel_path not in _EXCEL_HEADER_OK:
                try:
                    header = [str(c.value or "").strip() for c in ws[1]]
                    if "Trạng thái" not in header:
                        col = len(header) + 1
                        ws.cell(row=1, column=col, value="Trạng thái")
                        ws.cell(row=1, column=col).font = Font(bold=True)
                        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
                        # fill rỗng cho các dòng cũ
                        for r in range(2, ws.max_row + 1):
                            ws.cell(row=r, column=col, value=ws.cell(row=r, column=col).value or "")
                    _EXCEL_HEADER_OK.add(excel_path)
                except Exception:
                    pass
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "YouTube Links"
            ws.append(["STT", "Tên file", "Link YouTube", "Thời gian", "Trạng thái"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        next_index = ws.max_row  # header ở row 1 => dòng data đầu = 2 => STT = max_row
        ws.append([next_index, file_name or "N/A", url or "", datetime.now().strftime("%Y-%m-%d %H:%M"), status or ""])
        # Đảm bảo file cũ được "mở rộng" đủ 5 cột (một số trường hợp append không tăng max_column như mong đợi)
        try:
            ws.cell(row=1, column=5, value=ws.cell(row=1, column=5).value or "Trạng thái")
            ws.cell(row=ws.max_row, column=5, value=status or "")
        except Exception:
            pass
        if excel_path not in _EXCEL_WIDTHS_SET:
            _set_excel_column_widths(ws, 5)
            _EXCEL_WIDTHS_SET.add(excel_path)
        wb.save(excel_path)
        _log(log_callback, f"📄 Đã cập nhật Excel: {excel_path}")
        _dbg("E1", "append_excel_row saved", {"ms": int((time.time() - t0) * 1000), "file": os.path.basename(file_name or ""), "hasUrl": bool(url), "status": status, "excel": os.path.basename(excel_path)})
        try:
            _agent_debug_log(
                "X1",
                "append_excel_row saved",
                {"excelPath": excel_path, "row": ws.max_row, "maxCol": ws.max_column},
                run_id="excel_append_post",
            )
        except Exception:
            pass
        return excel_path
    except Exception as e:
        _log(log_callback, f"⚠️ Lỗi khi cập nhật Excel realtime: {e}")
        try:
            _agent_debug_log(
                "X2",
                "append_excel_row failed",
                {"excelPath": excel_path, "error": str(e)},
                run_id="excel_append_error",
            )
        except Exception:
            pass
        return None
# #endregion agent log


def ensure_excel_initialized(excel_filename: str = "YouTube_Upload_Links.xlsx", log_callback=None):
    """
    Đảm bảo file Excel tồn tại và có đủ header, đặc biệt cột "Trạng thái".
    Dùng ngay khi bắt đầu batch để không phụ thuộc việc upload thành công/thất bại.
    """
    if not OPENPYXL_AVAILABLE:
        _log(log_callback, "Thiếu thư viện openpyxl, không khởi tạo được Excel.")
        return None
    excel_path = _agent_excel_path(excel_filename)
    try:
        _agent_debug_log(
            "X0",
            "ensure_excel_initialized invoked",
            {"excelPath": excel_path, "exists": os.path.exists(excel_path), "cwd": os.getcwd()},
            run_id="excel_init_pre",
        )
    except Exception:
        pass
    try:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            if excel_path not in _EXCEL_HEADER_OK:
                header = [str(c.value or "").strip() for c in ws[1]]
                if "Trạng thái" not in header:
                    col = len(header) + 1
                    ws.cell(row=1, column=col, value="Trạng thái")
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col, value=ws.cell(row=r, column=col).value or "")
                _EXCEL_HEADER_OK.add(excel_path)
            if excel_path not in _EXCEL_WIDTHS_SET:
                _set_excel_column_widths(ws, 5)
                _EXCEL_WIDTHS_SET.add(excel_path)
            wb.save(excel_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "YouTube Links"
            ws.append(["STT", "Tên file", "Link YouTube", "Thời gian", "Trạng thái"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            _set_excel_column_widths(ws, 5)
            _EXCEL_WIDTHS_SET.add(excel_path)
            _EXCEL_HEADER_OK.add(excel_path)
            wb.save(excel_path)
        try:
            _agent_debug_log(
                "X0",
                "ensure_excel_initialized done",
                {"excelPath": excel_path},
                run_id="excel_init_post",
            )
        except Exception:
            pass
        return excel_path
    except Exception as e:
        _log(log_callback, f"⚠️ Lỗi khi khởi tạo Excel: {e}")
        try:
            _agent_debug_log(
                "X0",
                "ensure_excel_initialized failed",
                {"excelPath": excel_path, "error": str(e)},
                run_id="excel_init_error",
            )
        except Exception:
            pass
        return None


def _handle_prechecks_warning_after_done(driver, log_callback=None):
    """
    Sau khi bấm Lưu/Done, YouTube đôi khi hiện dialog cảnh báo pre-checks:
    - Nút: "Vẫn xuất bản" / "Quay lại"
    Điều kiện nhận biết thường thấy: banner khuyến nghị "Bạn nên giữ video này ở chế độ riêng tư..."
    """
    try:
        # #region agent log
        try:
            _agent_debug_log(
                "P1",
                "Checking prechecks warning banner/dialog after Done",
                {"url": driver.current_url},
                run_id="prechecks_check_pre",
            )
        except Exception:
            pass
        # #endregion agent log

        # 1) Kiểm tra banner khuyến nghị (nếu có)
        banner_present = False
        try:
            banner = driver.find_elements(By.CSS_SELECTOR, "ytcp-banner #message .subheading")
            for b in banner:
                txt = (b.text or "").strip()
                if "Bạn nên giữ video này ở chế độ riêng tư" in txt or "You should keep this video private" in txt:
                    banner_present = True
                    break
        except Exception:
            banner_present = False

        # 2) Nếu có dialog warning: ưu tiên bấm "Vẫn xuất bản"
        # Selector theo DOM bạn gửi: ytcp-prechecks-warning-dialog + secondary-action-button
        try:
            # implicit_wait=10s có thể làm WebDriverWait(2s) bị kéo thành ~10s.
            try:
                driver.implicitly_wait(0)
            except Exception:
                pass
            dialog = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "ytcp-prechecks-warning-dialog"))
            )
            try:
                driver.implicitly_wait(10)
            except Exception:
                pass
            try:
                _agent_debug_log(
                    "P1",
                    "Prechecks warning dialog detected",
                    {"bannerPresent": banner_present},
                    run_id="prechecks_check_post",
                )
            except Exception:
                pass

            publish_btn = None
            for sel in [
                "ytcp-prechecks-warning-dialog ytcp-button#secondary-action-button button",
                "ytcp-prechecks-warning-dialog button[aria-label*='Vẫn']",
                "ytcp-prechecks-warning-dialog button[aria-label*='xuất bản']",
                "ytcp-prechecks-warning-dialog button:has(.ytcpButtonShapeImpl__button-text-content)",
            ]:
                try:
                    btns = driver.find_elements(By.CSS_SELECTOR, sel)
                    if btns:
                        publish_btn = btns[0]
                        break
                except Exception:
                    continue
            if publish_btn is None:
                # XPath fallback by visible text
                try:
                    publish_btn = driver.find_element(By.XPATH, "//*[contains(text(),'Vẫn xuất bản') or contains(text(),'Vẫn xuất bản') or contains(text(),'Publish anyway')]/ancestor::button")
                except Exception:
                    publish_btn = None
            if publish_btn is not None:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", publish_btn)
                try:
                    publish_btn.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", publish_btn)
                # Đây là dialog prechecks có thể còn sót từ lần trước; log rõ ngữ cảnh để tránh hiểu nhầm.
                _log(log_callback, "Đã dọn cảnh báo pre-checks: Vẫn xuất bản.")
                try:
                    _agent_debug_log(
                        "P2",
                        "Clicked 'Vẫn xuất bản' on prechecks dialog",
                        {},
                        run_id="prechecks_action_post",
                    )
                except Exception:
                    pass
                return True
            return False
        except Exception:
            # Không có dialog
            try:
                driver.implicitly_wait(10)
            except Exception:
                pass
            try:
                _agent_debug_log(
                    "P1",
                    "No prechecks warning dialog detected",
                    {"bannerPresent": banner_present},
                    run_id="prechecks_check_post",
                )
            except Exception:
                pass
            return False
    except Exception as e:
        try:
            _agent_debug_log(
                "P0",
                "Error while handling prechecks warning",
                {"error": str(e)},
                run_id="prechecks_error",
            )
        except Exception:
            pass
        return False


def _click_next_js(driver):
    """Bấm Next bằng JS (shadow / nút con trong ytcp-button), khi Selenium click thường thất bại."""
    try:
        return bool(
            driver.execute_script(
                "try{"
                "var roots=[document.querySelector('ytcp-uploads-dialog'), document.body];"
                "function clickNb(nb){"
                "  if(!nb) return false;"
                "  var b=(nb.tagName==='YTCP-BUTTON')?(nb.querySelector('button')||nb):nb;"
                "  if(b.disabled||b.getAttribute('aria-disabled')==='true') return false;"
                "  (b.scrollIntoView?b.scrollIntoView({block:'center'}):0);"
                "  b.click(); return true;"
                "}"
                "for(var r=0;r<roots.length;r++){"
                "  var root=roots[r]||document;"
                "  var list=["
                "    root.querySelector&&root.querySelector('ytcp-button#next-button'),"
                "    root.querySelector&&root.querySelector('#next-button')"
                "  ].filter(Boolean);"
                "  for(var i=0;i<list.length;i++){ if(clickNb(list[i])) return true; }"
                "}"
                "var all=document.querySelectorAll('ytcp-button#next-button, #next-button');"
                "for(var j=0;j<all.length;j++){ if(clickNb(all[j])) return true; }"
                "return false;"
                "}catch(e){return false;}"
            )
        )
    except Exception:
        return False


def _wait_studio_next_enabled(driver, timeout_s=45, log_callback=None):
    """Chờ nút Next trong luồng upload bật (sau chọn trẻ em YouTube thường delay vài giây)."""
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        try:
            state = driver.execute_script(
                "try{"
                "var nb=document.querySelector('ytcp-uploads-dialog ytcp-button#next-button')"
                "||document.querySelector('ytcp-button#next-button')"
                "||document.querySelector('#next-button');"
                "if(!nb) return {ok:false, reason:'missing'};"
                "var b=(nb.tagName==='YTCP-BUTTON')?(nb.querySelector('button')||nb):nb;"
                "var dis=!!(b.disabled||b.getAttribute('aria-disabled')==='true');"
                "return {ok:!dis, reason: dis?'disabled':'ready'};"
                "}catch(e){return {ok:false, reason:'err'};}"
            )
            if state and state.get("ok"):
                return True
        except Exception:
            pass
        time.sleep(0.4)
    if log_callback:
        _log(log_callback, "⚠️ Nút Next vẫn chưa bật sau khi chờ — vẫn thử bấm (YouTube có thể chậm).")
    return False


def _click_next(driver, log_callback, step_name=""):
    """Bấm nút Next nếu có."""
    # Giả thuyết:
    # N1: Sau khi xử lý bản quyền và quay lại, nút Next dùng selector khác (#done-button, label tiếng Việt khác, v.v.)
    # N2: Có overlay/dialog che nút Next nên element_to_be_clickable không thỏa.
    # N3: Đang ở bước Chế độ hiển thị (chỉ có nút Lưu/Done, không còn Next).
    try:
        try:
            _agent_debug_log(
                "N1",
                "Trying to locate Next button",
                {"step": step_name},
                run_id="click_next_pre",
            )
        except Exception:
            pass
        # ytcp-button#next-button (bước Chi tiết, Các thành phần của video, Kiểm tra); aria-label tiếng Việt "Tiếp"
        # Tối ưu: thử nhanh trước (để không “đợi 10s” khi nút đã sẵn), fail mới fallback lâu hơn.
        sel = "ytcp-button#next-button, #next-button, button[aria-label='Tiếp'], button[aria-label='Next']"
        # Tắt implicit wait tạm thời để "thử nhanh 2s" thật sự là 2s (tránh bị kéo thành 10s).
        try:
            driver.implicitly_wait(0)
        except Exception:
            pass
        try:
            next_btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
        except TimeoutException:
            next_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
        finally:
            try:
                driver.implicitly_wait(10)
            except Exception:
                pass
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
        try:
            next_btn.click()
        except Exception:
            if not _click_next_js(driver):
                raise
        if step_name and log_callback:
            log_callback(f"Đã bấm Next: {step_name}")
        try:
            _agent_debug_log(
                "N1",
                "Clicked Next button successfully",
                {"step": step_name},
                run_id="click_next_post",
            )
        except Exception:
            pass
        return True
    except TimeoutException:
        if _click_next_js(driver):
            if step_name and log_callback:
                log_callback(f"Đã bấm Next (JS): {step_name}")
            try:
                _agent_debug_log(
                    "N1b",
                    "Clicked Next via JS fallback after timeout",
                    {"step": step_name},
                    run_id="click_next_js_ok",
                )
            except Exception:
                pass
            return True
        # Ghi log chi tiết khi không tìm thấy/click được Next
        try:
            _agent_debug_log(
                "N2",
                "Failed to locate/click Next button (TimeoutException)",
                {"step": step_name, "url": driver.current_url},
                run_id="click_next_error",
            )
        except Exception:
            pass
        return False


def _handle_checks_and_copyright(driver, wait, log_callback):
    """
    Bước Kiểm tra ban đầu (ytcp-uploads-checks): không báo gì thì Next.
    Nếu có bản quyền: xử lý thay thế bài hát rồi tiếp tục.
    :return: True nếu OK tiếp tục, False nếu không hoàn thành được (upload_video sẽ báo lỗi và chuyển video khác).
    """
    try:
        # Đọc nội dung mô tả kết quả kiểm tra bản quyền
        try:
            status_desc_el = driver.find_element(
                By.CSS_SELECTOR,
                "#copyright-status #results-description"
            )
            status_desc = (status_desc_el.text or "").strip()
        except Exception:
            status_desc_el = None
            status_desc = ""

        # Nếu hệ thống vẫn đang ở trạng thái "Kiểm tra xem video của bạn có chứa nội dung có bản quyền hay không"
        # thì chờ cho đến khi trạng thái đổi sang 1 trong 2:
        # - "Không phát hiện vấn đề nào"  -> tiếp tục luôn
        # - "Phát hiện có nội dung được bảo hộ bản quyền..." -> xử lý thay thế bài hát
        if status_desc and "Kiểm tra xem video của bạn có chứa nội dung có bản quyền hay không" in status_desc:
            _log(log_callback, "Đang chờ YouTube chạy kiểm tra bản quyền (tối đa ~3 phút)...")
            max_wait_secs = 180
            waited = 0
            while waited < max_wait_secs:
                time.sleep(3)
                waited += 3
                try:
                    status_desc_el = driver.find_element(
                        By.CSS_SELECTOR,
                        "#copyright-status #results-description"
                    )
                    status_desc = (status_desc_el.text or "").strip()
                except Exception:
                    status_desc = ""
                    break
                # Khi text không còn là câu "Kiểm tra xem video..." nữa thì thoát vòng lặp để xử lý tiếp
                if "Kiểm tra xem video của bạn có chứa nội dung có bản quyền hay không" not in status_desc:
                    break
            _log(log_callback, f"Trạng thái kiểm tra bản quyền sau khi chờ: {status_desc or 'không xác định'}")

        # Trạng thái đã xong: không phát hiện vấn đề nào -> bỏ qua bước bản quyền, Next luôn
        if status_desc and "Không phát hiện vấn đề nào" in status_desc:
            _log(log_callback, "Kiểm tra bản quyền: Không phát hiện vấn đề nào, bỏ qua bước Xem chi tiết.")
            # Nút Next ở bước này thường đã sẵn sàng ngay khi ra kết quả.
            # Thử bấm ngay bằng JS để tránh delay do implicit/overlay; nếu fail thì luồng ngoài sẽ bấm tiếp.
            try:
                clicked_next = driver.execute_script(
                    "try{"
                    "var cand=["
                    "document.querySelector('ytcp-uploads-dialog ytcp-button#next-button'),"
                    "document.querySelector('ytcp-button#next-button'),"
                    "document.querySelector('#next-button'),"
                    "document.querySelector(\"button[aria-label*='Tiếp'], button[aria-label*='Next']\")"
                    "].filter(Boolean);"
                    "for(var i=0;i<cand.length;i++){"
                    "  var el=cand[i];"
                    "  var b=(el.tagName==='YTCP-BUTTON')?(el.querySelector('button')||el):el;"
                    "  var dis=b.disabled || b.getAttribute('aria-disabled')==='true' || b.hasAttribute('disabled');"
                    "  if(!dis){"
                    "    (b.scrollIntoView?b.scrollIntoView({block:'center'}):0);"
                    "    b.click();"
                    "    return true;"
                    "  }"
                    "}"
                    "return false;"
                    "}catch(e){return false;}"
                )
                if clicked_next:
                    _log(log_callback, "Đã bấm Next ngay khi có kết quả kiểm tra bản quyền.")
            except Exception:
                pass
            return True

        # Theo yêu cầu: chỉ cần có dòng "Phát hiện có nội dung được bảo hộ bản quyền"
        # thì luôn ưu tiên bấm "Xem chi tiết" để chạy các bước xử lý tiếp theo.

        # Có claim bản quyền -> luôn mở "Xem chi tiết"
        if status_desc and "Phát hiện có nội dung được bảo hộ bản quyền" in status_desc:
            _log(log_callback, "Phát hiện cảnh báo bản quyền, đang chờ nút Xem chi tiết (YouTube kiểm tra có thể vài phút)...")
            # Chờ nút "Xem chi tiết" xuất hiện — tối đa 5 phút, không báo lỗi trong lúc chờ
            #
            # Lưu ý: UI mới nhiều khi render là <button class="ytcpButtonShapeImplHost"... aria-label="Xem chi tiết">
            # (không phải ytcp-button), nên ưu tiên bắt theo aria-label/text và JS-click.
            clicked = False
            try:
                # JS-fast probe/click (ổn định với cả button shape mới)
                t0 = time.time()
                # Thử click trong một khoảng hợp lý; nếu YouTube vẫn render chậm thì fallback bằng wait khác.
                while (time.time() - t0) < 120:
                    ok = False
                    try:
                        ok = bool(driver.execute_script(
                            "try{"
                            "var root=document.querySelector('#copyright-status')||document;"
                            "var b=root.querySelector(\"button[aria-label*='Xem chi'],button[aria-label*='View details'],button[title*='Xem chi']\");"
                            "if(!b){"
                            "  var all=[...root.querySelectorAll('button')];"
                            "  b=all.find(x=>((x.innerText||'').trim().toLowerCase().includes('xem chi')||(x.innerText||'').trim().toLowerCase().includes('view details')));"
                            "}"
                            "if(!b) return false;"
                            "var dis=b.disabled||b.getAttribute('aria-disabled')==='true'||b.hasAttribute('disabled');"
                            "if(dis) return false;"
                            "b.scrollIntoView({block:'center'});"
                            "b.click();"
                            "return true;"
                            "}catch(e){return false;}"
                        ))
                    except Exception:
                        ok = False
                    if ok:
                        clicked = True
                        _log(log_callback, "Đã bấm Xem chi tiết.")
                        time.sleep(3)
                        break
                    time.sleep(1.0)
            except Exception:
                pass
            if not clicked:
                try:
                    btn = WebDriverWait(driver, 120).until(EC.presence_of_element_located((
                        By.XPATH,
                        "//*[self::button or self::ytcp-button or self::tp-yt-paper-button][contains(@aria-label,'Xem chi') or contains(.,'Xem chi') or contains(.,'View details')]"
                    )))
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                        driver.execute_script("arguments[0].click();", btn)
                    except Exception:
                        btn.click()
                    clicked = True
                    _log(log_callback, "Đã bấm Xem chi tiết.")
                    time.sleep(3)
                except Exception:
                    pass
            if not clicked:
                _log(log_callback, "❌ Lỗi: Có cảnh báo bản quyền nhưng không bấm được nút Xem chi tiết. Chuyển video khác.")
                return False
            time.sleep(2)
            # Chọn cách giải quyết — chờ dialog mở rồi mới bấm (tối đa 15s)
            choose_clicked = False
            try:
                choose_btn = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "ytcr-video-actions-button button[aria-label='Chọn cách giải quyết'], "
                    ".resolution-actions-button ytcp-button, "
                    "ytcr-video-actions-button #actions-button, "
                    "button[aria-label='Chọn cách giải quyết']"
                )))
                choose_btn.click()
                choose_clicked = True
                _log(log_callback, "Đã bấm Chọn cách giải quyết.")
                time.sleep(2)
            except Exception:
                pass
            if not choose_clicked:
                for btn_text in ["Chọn cách giải quyết", "Choose resolution", "Choose action"]:
                    try:
                        btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((
                            By.XPATH, f"//*[contains(text(),'{btn_text}')]"
                        )))
                        btn.click()
                        time.sleep(2)
                        choose_clicked = True
                        break
                    except Exception:
                        continue
            if not choose_clicked:
                _log(log_callback, "❌ Lỗi: Không tìm thấy Chọn cách giải quyết. Chuyển video khác.")
                return False
            time.sleep(2)
            # Thay thế bài hát — trong ytcr-video-actions-dialog: action-card-container action="NON_TAKEDOWN_CLAIM_OPTION_REPLACE_SONG"
            replace_clicked = False
            try:
                replace_btn = driver.find_element(
                    By.CSS_SELECTOR,
                    "ytcr-video-actions-dialog button.action-card-container[action='NON_TAKEDOWN_CLAIM_OPTION_REPLACE_SONG'], "
                    "ytcr-video-actions-dialog .action-card-container[action='NON_TAKEDOWN_CLAIM_OPTION_REPLACE_SONG'], "
                    "[action='NON_TAKEDOWN_CLAIM_OPTION_REPLACE_SONG']"
                )
                replace_btn.click()
                replace_clicked = True
                _log(log_callback, "Đã chọn Thay thế bài hát.")
                time.sleep(2)
            except Exception:
                pass
            if not replace_clicked:
                for opt in ["Thay thế bài hát", "Replace the song", "Replace song"]:
                    try:
                        el = driver.find_element(By.XPATH, f"//*[contains(text(),'{opt}')]")
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                        try:
                            el.click()
                        except Exception:
                            driver.execute_script("arguments[0].click();", el)
                        replace_clicked = True
                        _log(log_callback, "Đã chọn Thay thế bài hát (theo text).")
                        time.sleep(2)
                        break
                    except Exception:
                        continue
            if not replace_clicked:
                try:
                    ok = bool(
                        driver.execute_script(
                            "try{"
                            "var dlg=document.querySelector('ytcr-video-actions-dialog')||document;"
                            "var cards=dlg.querySelectorAll('[action=\\'NON_TAKEDOWN_CLAIM_OPTION_REPLACE_SONG\\'], .action-card-container');"
                            "for(var i=0;i<cards.length;i++){"
                            "  var t=(cards[i].innerText||'').toLowerCase();"
                            "  if(t.includes('thay thế')||t.includes('replace')){"
                            "    cards[i].scrollIntoView({block:'center'}); cards[i].click(); return true;"
                            "  }"
                            "}"
                            "return false;"
                            "}catch(e){return false;}"
                        )
                    )
                    if ok:
                        replace_clicked = True
                        _log(log_callback, "Đã chọn Thay thế bài hát (JS).")
                        time.sleep(2)
                except Exception:
                    pass
            time.sleep(1)
            # Trong dialog sau "Thay thế bài hát" có nút "Tiếp tục" (confirm-button) — bấm để mở panel chọn bài hát
            confirm_clicked = False
            try:
                confirm_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "ytcr-video-actions-dialog ytcp-button#confirm-button, "
                    "ytcr-video-actions-dialog button[aria-label*='Tiếp tục'], "
                    "ytcr-video-actions-dialog button[aria-label*='Tiếp tục'], "
                    "button[aria-label='Tiếp tục']"
                )))
                try:
                    inner = confirm_btn.find_element(By.CSS_SELECTOR, "button")
                    inner.click()
                except Exception:
                    confirm_btn.click()
                confirm_clicked = True
                _log(log_callback, "Đã bấm Tiếp tục trong dialog thay thế bài hát.")
                time.sleep(2)
            except Exception:
                pass
            if not confirm_clicked:
                try:
                    btn = driver.find_element(By.XPATH, "//*[contains(text(),'Tiếp tục') or contains(text(),'Continue')]/ancestor::button | //ytcp-button[.//*[contains(text(),'Tiếp tục')]]")
                    btn.click()
                    confirm_clicked = True
                    _log(log_callback, "Đã bấm Tiếp tục.")
                    time.sleep(2)
                except Exception:
                    pass
            # Chờ panel thay thế bài hát tải xong
            time.sleep(3)
            _log(log_callback, "Tiếp tục: chọn bài hát đầu tiên (bấm Thêm)...")
            # Nút Thêm: ytcp-icon-button#add-track-button (aria-label="Thêm") trong ytve-audioswap-track-row — bấm cái đầu tiên
            add_clicked = False
            try:
                first_add = WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located((
                    By.CSS_SELECTOR,
                    "ytve-audioswap-track-row ytcp-icon-button#add-track-button, "
                    "ytcp-icon-button#add-track-button, "
                    "ytcp-icon-button[aria-label='Thêm']"
                )))
                if first_add:
                    add_el = first_add[0]
                    time.sleep(0.5)
                    add_el.click()
                    add_clicked = True
                    _log(log_callback, "Đã chọn bài hát đầu tiên (Thêm).")
                    time.sleep(2)
            except Exception:
                pass
            if not add_clicked:
                try:
                    plus_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
                        By.CSS_SELECTOR,
                        "ytcp-icon-button#add-track-button, ytcp-icon-button[aria-label='Thêm']"
                    )))
                    plus_btn.click()
                    add_clicked = True
                    _log(log_callback, "Đã chọn bài hát đầu tiên (Thêm).")
                    time.sleep(2)
                except Exception:
                    plus_btns = driver.find_elements(By.CSS_SELECTOR, "ytcp-icon-button#add-track-button, ytcp-icon-button[aria-label='Thêm']")
                    if plus_btns:
                        plus_btns[0].click()
                        add_clicked = True
                        _log(log_callback, "Đã chọn bài hát đầu tiên (Thêm).")
                        time.sleep(2)
            time.sleep(2)
            # Lưu / Xong — chờ nút sẵn sàng (tối đa 15s), thử click nút thật bên trong ytcp-button nếu cần
            _log(log_callback, "Đang bấm Lưu / Xong...")
            save_clicked = False
            # Ưu tiên nút Lưu trên góc phải trong modal "Thay thế bài hát"
            try:
                # Tìm đúng button bên trong ytcp-button#save-button như DOM bạn gửi
                save_btn_outer = WebDriverWait(driver, 15).until(EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "ytve-modal-host ytcp-button#save-button"
                )))
                inner_button = save_btn_outer.find_element(By.CSS_SELECTOR, "button[aria-label='Lưu'], button.ytcpButtonShapeImplHost")
                # Đảm bảo nút ở trong vùng hiển thị và không bị lớp khác che
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", inner_button)
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(inner_button))
                try:
                    inner_button.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", inner_button)
                save_clicked = True
                _log(log_callback, "Đã bấm nút Lưu trong modal Thay thế bài hát.")
                time.sleep(2)
            except Exception:
                pass
            # Nếu chưa bấm được Lưu, fallback sang Xong/Done trong editor
            if not save_clicked:
                try:
                    done_btn = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((
                        By.CSS_SELECTOR,
                        "ytmus-license-dialog ytcp-button#done-button, "
                        "ytve-editor ytcp-button#done-button, "
                        "button[aria-label='Xong'], button[aria-label='Done']"
                    )))
                    try:
                        inner = done_btn.find_element(By.CSS_SELECTOR, "button")
                        inner.click()
                    except Exception:
                        done_btn.click()
                    save_clicked = True
                    _log(log_callback, "Đã bấm Xong.")
                    time.sleep(2)
                except Exception:
                    pass
            if not save_clicked:
                for save_text in ["Lưu", "Xong", "Save", "Lưu thay đổi", "Done"]:
                    try:
                        save_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                            By.XPATH, f"//*[contains(text(),'{save_text}')]/ancestor::button | //button[contains(.,'{save_text}')]"
                        )))
                        save_btn.click()
                        save_clicked = True
                        _log(log_callback, f"Đã bấm {save_text}.")
                        time.sleep(2)
                        break
                    except Exception:
                        continue
            if not save_clicked:
                _log(log_callback, "❌ Lỗi: Không tìm thấy nút Lưu/Xong sau thay thế bài hát. Chuyển video khác.")
                return False

            # Hộp thoại "Xác nhận thay đổi": tích checkbox trước, chờ nút Xác nhận bật rồi bấm
            time.sleep(2)
            try:
                # Tìm đúng checkbox theo DOM bạn gửi
                checkbox = WebDriverWait(driver, 12).until(
                    EC.presence_of_element_located((
                        By.CSS_SELECTOR,
                        "ytve-save-dialog ytcp-checkbox-lit #checkbox, "
                        "ytcp-checkbox-lit #checkbox[role='checkbox'][aria-label*='xác nhận'], "
                        "div#checkbox[role='checkbox'][aria-label*='xác nhận']"
                    ))
                )
                # Đưa checkbox vào giữa màn hình và bấm bằng JS để tránh lớp che
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                WebDriverWait(driver, 8).until(EC.element_to_be_clickable(checkbox))
                try:
                    checkbox.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", checkbox)
                # Nếu vẫn chưa được tick thì gửi phím Space
                time.sleep(0.8)
                aria_checked = checkbox.get_attribute("aria-checked")
                if aria_checked not in ("true", "True"):
                    from selenium.webdriver.common.keys import Keys as _Keys
                    checkbox.send_keys(_Keys.SPACE)
                    time.sleep(0.8)
                aria_checked = checkbox.get_attribute("aria-checked")
                _log(log_callback, f"Trạng thái checkbox sau khi click: aria-checked={aria_checked}")
                if aria_checked not in ("true", "True"):
                    _log(log_callback, "⚠️ Không tick được checkbox xác nhận thay đổi vĩnh viễn, vẫn thử bấm nút 'Xác nhận thay đổi'.")
                else:
                    _log(log_callback, "Đã tích xác nhận thay đổi vĩnh viễn.")
            except Exception as e:
                _log(log_callback, f"⚠️ Không tìm được checkbox 'Tôi xác nhận rằng những thay đổi này là vĩnh viễn': {e}")

            try:
                # Tìm đúng nút 'Xác nhận thay đổi' theo DOM bạn gửi
                apply_btn_outer = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((
                        By.CSS_SELECTOR,
                        "ytve-save-dialog ytcp-button#apply-button"
                    ))
                )
                apply_inner = apply_btn_outer.find_element(
                    By.CSS_SELECTOR,
                    "button[aria-label='Xác nhận thay đổi'], button.ytcpButtonShapeImplHost"
                )
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", apply_inner)
                WebDriverWait(driver, 8).until(EC.element_to_be_clickable(apply_inner))
                try:
                    apply_inner.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", apply_inner)
                _log(log_callback, "Đã bấm Xác nhận thay đổi.")
                time.sleep(3)
            except Exception:
                try:
                    apply_btn = driver.find_element(By.CSS_SELECTOR, "ytcp-button#apply-button")
                    inner = apply_btn.find_element(By.CSS_SELECTOR, "button")
                    if inner.get_attribute("disabled") != "true":
                        inner.click()
                        _log(log_callback, "Đã bấm Xác nhận thay đổi (fallback).")
                        time.sleep(3)
                except Exception:
                    for txt in ["Xác nhận thay đổi", "Apply changes", "Xác nhận"]:
                        try:
                            btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                                By.XPATH, f"//button[contains(.,'{txt}')] | //ytcp-button[.//*[contains(text(),'{txt}')]]"
                            )))
                            btn.click()
                            _log(log_callback, f"Đã bấm '{txt}'.")
                            time.sleep(3)
                            break
                        except Exception:
                            continue

            # Sau khi áp dụng thay đổi, quay lại editor / quy trình tải lên
            try:
                for back_txt in ["Quay lại quy trình tải lên", "Quay lại trình chỉnh sửa", "Quay lại"]:
                    try:
                        back_btn = driver.find_element(
                            By.XPATH,
                            f"//ytcp-button[.//*[contains(text(),'{back_txt}')]] | //button[contains(text(),'{back_txt}')]"
                        )
                        back_btn.click()
                        _log(log_callback, f"Đã bấm nút '{back_txt}' để quay lại quy trình tải lên.")
                        time.sleep(2)
                        break
                    except Exception:
                        continue
            except Exception:
                pass

            # Sau khi quay lại quy trình tải lên, bấm "Tiếp" để sang bước "Chế độ hiển thị"
            try:
                if _click_next(driver, log_callback, "Sau thay thế bài hát → Chế độ hiển thị"):
                    time.sleep(2)
            except Exception:
                pass

            _log(log_callback, "Đã xử lý thay thế bài hát và xác nhận thay đổi, quay lại quy trình tải lên.")
            return True
        else:
            _log(log_callback, "Không có cảnh báo kiểm tra, tiếp tục.")
            return True
    except Exception as e:
        _log(log_callback, f"❌ Lỗi bước kiểm tra/bản quyền: {e}. Chuyển video khác.")
        return False


def ensure_youtube_login(driver, email, password, log_callback=None):
    """
    Nếu Chrome đang ở trang đăng nhập Google/YouTube thì tự động điền email + mật khẩu và đăng nhập.
    Gọi sau init_driver khi profile đã lưu email/password. Nếu có 2FA/captcha thì chỉ ghi log, user đăng nhập tay.
    """
    if not email or not str(email).strip() or not password:
        return
    email = str(email).strip()
    try:
        driver.get(YOUTUBE_STUDIO_URL)
        time.sleep(3)
        url = driver.current_url
        _dbg("LG1", "ensure_youtube_login landing", {"url": (url or "")[:160]})
        if "accounts.google.com" in url or "signin" in url.lower() or "login" in url.lower():
            _log(log_callback, "Đang tự động đăng nhập bằng email/mật khẩu đã lưu...")
        else:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            _dbg("LG2", "ensure_youtube_login body scan", {"hasSignInText": ("Đăng nhập" in body_text) or ("Sign in" in body_text) or ("Sign in to YouTube" in body_text), "url": (driver.current_url or "")[:160]})
            if "Đăng nhập" in body_text or "Sign in" in body_text or "Sign in to YouTube" in body_text:
                _log(log_callback, "Phát hiện trang đăng nhập, đang điền email/mật khẩu...")
            else:
                _log(log_callback, "Đã đăng nhập sẵn (cookie profile).")
                _dbg("LG3", "ensure_youtube_login already logged", {"url": (driver.current_url or "")[:160]})
                return
        if "accounts.google.com" not in driver.current_url:
            try:
                signin = driver.find_element(By.XPATH, "//a[contains(@href,'accounts.google.com') or contains(text(),'Đăng nhập') or contains(text(),'Sign in')]")
                signin.click()
                time.sleep(3)
            except Exception:
                driver.get("https://accounts.google.com/ServiceLogin?service=youtube&continue=https://studio.youtube.com/")
                time.sleep(3)
        try:
            email_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "input[type='email'], input[name='identifier'], #identifierId"
            )))
            email_input.clear()
            email_input.send_keys(email)
            _log(log_callback, "Đã nhập email.")
            time.sleep(0.5)
        except Exception as e:
            _log(log_callback, f"Không tìm thấy ô email: {e}")
            return
        try:
            next_btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((
                By.XPATH,
                "//span[text()='Next']/.. | //span[text()='Tiếp']/.. | //button[.//span[text()='Next']] | //div[@role='button']//span[text()='Tiếp']/.. | //*[text()='Next']/ancestor::button | //*[text()='Tiếp']/ancestor::div[@role='button']"
            )))
            next_btn.click()
            time.sleep(3)
        except Exception:
            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, "#identifierNext button, [data-idom-class*='Next'] button")
                next_btn.click()
                time.sleep(3)
            except Exception as e:
                _log(log_callback, f"Không bấm được Next: {e}")
                return
        try:
            pw_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "input[type='password'], input[name='password'], input[name='Passwd']"
            )))
            pw_input.clear()
            pw_input.send_keys(password)
            _log(log_callback, "Đã nhập mật khẩu.")
            time.sleep(0.5)
        except Exception as e:
            _log(log_callback, f"Không tìm thấy ô mật khẩu (có thể cần 2FA): {e}")
            return
        try:
            next_btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((
                By.XPATH,
                "//span[text()='Next']/.. | //span[text()='Tiếp']/.. | //button[.//span[text()='Next']] | //*[text()='Tiếp']/ancestor::div[@role='button']"
            )))
            next_btn.click()
            time.sleep(5)
        except Exception:
            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, "#passwordNext button, button[type='submit']")
                next_btn.click()
                time.sleep(5)
            except Exception:
                pass
        time.sleep(3)
        if "accounts.google.com" in driver.current_url:
            _log(log_callback, "Có thể cần xác minh 2 bước hoặc captcha — vui lòng đăng nhập tay trên Chrome.")
            _dbg("LG9", "ensure_youtube_login still on accounts.google.com", {"url": (driver.current_url or "")[:160]})
        else:
            _log(log_callback, "Đã đăng nhập xong (tự động).")
            _dbg("LG8", "ensure_youtube_login done", {"url": (driver.current_url or "")[:160]})
    except Exception as e:
        _log(log_callback, f"Tự động đăng nhập lỗi: {e}")
        _dbg("LGX", "ensure_youtube_login exception", {"err": str(e)[:200]})


def init_driver(headless=False, use_saved_profile=True, profile_dir=None):
    """
    Khởi tạo Chrome WebDriver để dùng cho upload.
    :param headless: Chạy ẩn browser (không nên dùng vì cần đăng nhập thủ công).
    :return: WebDriver instance hoặc None nếu lỗi.
    """
    # Giả thuyết:
    # H1: Không tìm thấy Chrome hoặc ChromeDriver -> WebDriverException ở webdriver.Chrome
    # H2: Lỗi khi tạo/thao tác thư mục profile (quyền, đường dẫn sai)
    # H3: Fallback Edge cũng lỗi do không có Edge hoặc driver tương ứng
    try:
        options = Options()
        if headless:
            options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        # Giảm lỗi timeout khi user mở Task Manager, cửa sổ khác che Chrome hoặc tải file nặng
        options.add_argument("--disable-backgrounding-occluded-windows")
        options.add_argument("--disable-renderer-backgrounding")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        # Giữ trình duyệt mở để user có thể đăng nhập
        options.add_experimental_option("detach", False)
        # Cho phép driver.get_log("browser") — dùng khi debug lỗi JS/Studio
        try:
            options.set_capability("goog:loggingPrefs", {"browser": "ALL"})
        except Exception:
            pass

        chosen_profile_dir = None
        if use_saved_profile:
            chosen_profile_dir = profile_dir or CHROME_PROFILE_DIR
            _agent_debug_log(
                "H2",
                "About to create/use Chrome profile directory",
                {"profileDir": chosen_profile_dir},
            )
            os.makedirs(chosen_profile_dir, exist_ok=True)
            options.add_argument(f"--user-data-dir={chosen_profile_dir}")
            options.add_argument("--profile-directory=Default")

            # #region agent log (debug-mode NDJSON)
            try:
                _dbg("CH1", "init_driver profile prepared", {"profileDir": chosen_profile_dir})
            except Exception:
                pass
            # #endregion agent log

            # Xóa file DevToolsActivePort còn sót lại (thường gây session not created / invalid session)
            try:
                dtap = os.path.join(chosen_profile_dir, "DevToolsActivePort")
                if os.path.exists(dtap):
                    try:
                        os.remove(dtap)
                        _dbg("CH2", "removed DevToolsActivePort", {"path": "DevToolsActivePort"})
                    except Exception as e:
                        _dbg("CH2", "failed remove DevToolsActivePort", {"err": str(e)[:120]})
            except Exception:
                pass

        _agent_debug_log(
            "H1",
            "Attempting to start Chrome WebDriver",
            {"useSavedProfile": use_saved_profile, "profileDir": chosen_profile_dir},
        )
        # Thử dùng Chrome mặc định (chromedriver trong PATH hoặc Selenium Manager)
        try:
            driver = webdriver.Chrome(options=options)
        except WebDriverException as e:
            # #region agent log (debug-mode NDJSON)
            try:
                _dbg("CHX", "webdriver.Chrome failed (first attempt)", {"err": str(e)[:200]})
            except Exception:
                pass
            # #endregion agent log
            # Retry 1 lần (sau khi đã xoá DevToolsActivePort)
            driver = webdriver.Chrome(options=options)
        driver.maximize_window()
        driver.implicitly_wait(10)
        # Giảm nguy cơ execute_script bị treo lâu (đặc biệt khi Studio lag/overlay).
        try:
            driver.set_script_timeout(6)
        except Exception:
            pass
        try:
            driver.set_page_load_timeout(45)
        except Exception:
            pass
        try:
            _dbg("CH3", "driver timeouts set", {"scriptTimeoutS": 6, "pageLoadTimeoutS": 45})
        except Exception:
            pass
        _agent_debug_log(
            "H1",
            "Chrome WebDriver started successfully",
            {"useSavedProfile": use_saved_profile},
            run_id="init_driver_post",
        )
        return driver
    except WebDriverException as e:
        # Chỉ dùng Chrome, không fallback Edge
        _agent_debug_log(
            "H1",
            "Chrome WebDriver failed with WebDriverException (no Edge fallback)",
            {"error": str(e)},
            run_id="init_driver_error",
        )
        # #region agent log (debug-mode NDJSON)
        try:
            _dbg("CH9", "init_driver WebDriverException bubbled", {"err": str(e)[:220]})
        except Exception:
            pass
        # #endregion agent log
        # Ném lại lỗi để phía trên ghi log rõ ràng, tránh tự động chuyển sang Edge
        raise e


def _try_get_video_link_from_page(driver):
    """Thử lấy link video từ trang hiện tại (trước hoặc sau khi bấm Done). Trả về url hoặc None."""
    try:
        def _extract_video_id(u: str):
            if not u:
                return None
            u = str(u).strip()
            # watch?v=
            m = re.search(r"[?&]v=([\w-]{6,})", u)
            if m:
                return m.group(1)
            # youtu.be/<id>
            m = re.search(r"youtu\.be/([\w-]{6,})", u)
            if m:
                return m.group(1)
            # youtube.com/shorts/<id>
            m = re.search(r"youtube\.com/shorts/([\w-]{6,})", u)
            if m:
                return m.group(1)
            # studio.youtube.com/video/<id>/...
            m = re.search(r"studio\.youtube\.com/video/([\w-]{6,})", u)
            if m:
                return m.group(1)
            return None

        def _normalize_youtube_url(u: str):
            vid = _extract_video_id(u)
            if vid:
                return f"https://youtu.be/{vid}"
            return None

        link_el = driver.find_elements(
            By.CSS_SELECTOR,
            "ytcp-video-info a[href*='youtu.be'], "
            "ytcp-video-info a[href*='youtube.com/watch'], "
            "ytcp-video-info a[href*='youtube.com/shorts'], "
            "ytcp-video-info a[href*='studio.youtube.com/video']"
        )
        if not link_el:
            link_el = driver.find_elements(
                By.CSS_SELECTOR,
                "a[href*='youtu.be'], a[href*='youtube.com/watch'], a[href*='youtube.com/shorts'], a[href*='studio.youtube.com/video']"
            )
        if link_el:
            url = link_el[0].get_attribute("href")
            if url:
                n = _normalize_youtube_url(url)
                if n:
                    return n
                return url.split("&")[0].strip()
        page_source = driver.page_source
        for pattern in (
            r"https?://(?:www\.)?youtu\.be/[\w-]+",
            r"https?://(?:www\.)?youtube\.com/watch\?v=[\w-]+",
            r"https?://(?:www\.)?youtube\.com/shorts/[\w-]+",
            r"https?://studio\.youtube\.com/video/[\w-]+[^\s\"']*",
        ):
            match = re.search(pattern, page_source)
            if match:
                u = match.group(0).split("&")[0]
                n = _normalize_youtube_url(u)
                return n or u
    except Exception:
        pass
    return None


def _dismiss_blocking_dialogs(driver, log_callback=None):
    """Đóng các dialog/overlay hay chặn thao tác (không ghi dữ liệu nhạy cảm)."""
    try:
        t0 = time.time()
        _dbg("Ddlg0", "dismiss dialogs enter", {"url": (getattr(driver, "current_url", "") or "")[:160]})
        # Budget nhỏ để tránh treo lâu ở bước dismiss (log runtime cho thấy có thể tốn 60-70s).
        max_budget_s = 6.0
        # JS-first: tránh các Selenium find_elements bị stall lâu (runtime evidence).
        try:
            driver.execute_script(
                "try{"
                "document.querySelectorAll('tp-yt-paper-dialog').forEach(d=>{d.style.display='none'; d.setAttribute('aria-hidden','true');});"
                "document.querySelectorAll('.tp-yt-iron-overlay-backdrop').forEach(b=>{b.style.display='none';});"
                "}catch(e){}"
            )
        except Exception:
            pass
        # prechecks warning (nếu còn)
        try:
            has_prechecks = bool(driver.execute_script("try{return !!document.querySelector('ytcp-prechecks-warning-dialog');}catch(e){return false;}"))
            if has_prechecks:
                _dbg("Ddlg1", "prechecks dialog present", {})
                # JS-fast: dọn popup prechecks còn sót (ưu tiên selector ổn định).
                try:
                    clicked = driver.execute_script(
                        "try{"
                        "var d=document.querySelector('ytcp-prechecks-warning-dialog');"
                        "if(!d) return false;"
                        "var b = d.querySelector(\"ytcp-button#secondary-action-button button, #secondary-action-button button, button[aria-label*='Vẫn'], button[aria-label*='xuất bản']\");"
                        "if(b){b.click(); return true;}"
                        "// fallback by text"
                        "var btns=[...d.querySelectorAll('button')];"
                        "var texts=['Vẫn xuất bản','Publish anyway','Xuất bản','Publish','OK','Đồng ý','Accept','Got it'];"
                        "for(var x of btns){var t=(x.innerText||'').trim();"
                        "for(var s of texts){if(t && t.includes(s)){x.click(); return true;}}}"
                        "return false;"
                        "}catch(e){return false;}"
                    )
                    _dbg("Ddlg1j", "prechecks js-click attempted", {"clicked": bool(clicked)})
                except Exception:
                    pass
                # JS-fast: remove dialog/backdrop để tránh “dính” sang video sau
                try:
                    driver.execute_script(
                        "try{"
                        "document.querySelectorAll('ytcp-prechecks-warning-dialog').forEach(x=>x.remove());"
                        "document.querySelectorAll('.tp-yt-iron-overlay-backdrop').forEach(b=>b.remove());"
                        "}catch(e){}"
                    )
                except Exception:
                    pass
                # Không gọi handler nặng ở đây để tránh stall ~20s; handler được dùng ở bước sau Done.
        except Exception:
            pass

        # Uploads dialog: thử bấm nút close (X) hoặc ESC
        try:
            has_uploads = bool(driver.execute_script("try{return !!document.querySelector('ytcp-uploads-dialog');}catch(e){return false;}"))
            if has_uploads:
                _dbg("Ddlg2", "uploads dialog present", {})
                # JS-first: đóng + remove ngay, tránh find_elements/Wait có thể stall lâu
                try:
                    driver.execute_script(
                        "try{"
                        "var d=document.querySelector('ytcp-uploads-dialog');"
                        "if(d){"
                        "var c=d.querySelector('#close-button, ytcp-icon-button#close-button, button[aria-label*=\\'Đóng\\'], button[aria-label*=\\'Close\\']');"
                        "if(c){c.click();}"
                        "d.style.display='none'; d.setAttribute('aria-hidden','true');"
                        "}"
                        "document.querySelectorAll('ytcp-uploads-dialog').forEach(x=>x.remove());"
                        "document.querySelectorAll('.tp-yt-iron-overlay-backdrop').forEach(b=>{b.style.display='none';});"
                        "document.querySelectorAll('.tp-yt-iron-overlay-backdrop').forEach(b=>b.remove());"
                        "}catch(e){}"
                    )
                    _dbg("Ddlg2j", "uploads js-close/remove applied", {})
                except Exception:
                    pass
        except Exception:
            pass

        # Generic tp-yt-paper-dialog (cookie/confirm): bấm nút đóng nếu có, hoặc ESC
        try:
            has_tp = bool(driver.execute_script("try{return !!document.querySelector('tp-yt-paper-dialog');}catch(e){return false;}"))
            if has_tp:
                _dbg("Ddlg3", "tp dialog present", {})
                # Runtime evidence: các bước find/click có thể stall ~20s.
                # Với tp-yt-paper-dialog (cookie/confirm), ưu tiên remove/hide thẳng bằng JS để không chặn click.
                try:
                    _dbg("Ddlg6", "tp dialog force hide/remove (js-fast)", {})
                except Exception:
                    pass
                try:
                    driver.execute_script(
                        "try{"
                        "document.querySelectorAll('tp-yt-paper-dialog').forEach(d=>{d.style.display='none'; d.setAttribute('aria-hidden','true');});"
                        "document.querySelectorAll('.tp-yt-iron-overlay-backdrop').forEach(b=>{b.style.display='none';});"
                        "document.querySelectorAll('tp-yt-paper-dialog').forEach(d=>d.remove());"
                        "document.querySelectorAll('.tp-yt-iron-overlay-backdrop').forEach(b=>b.remove());"
                        "}catch(e){}"
                    )
                except Exception:
                    pass
                # chờ dialog biến mất (timeout ngắn để tránh bị treo driver)
                try:
                    _dbg("Ddlg3w", "wait tp dialog gone (begin)", {})
                except Exception:
                    pass
                try:
                    # Quan trọng: không dùng find_elements ở đây vì implicit_wait có thể kéo dài thành ~10s.
                    # Dùng JS-probe + tắt implicit trong lúc wait.
                    try:
                        driver.implicitly_wait(0)
                    except Exception:
                        pass
                    try:
                        WebDriverWait(driver, 2).until(
                            lambda d: not bool(
                                d.execute_script(
                                    "try{return !!document.querySelector('tp-yt-paper-dialog');}catch(e){return false;}"
                                )
                            )
                        )
                    finally:
                        try:
                            driver.implicitly_wait(10)
                        except Exception:
                            pass
                except Exception:
                    pass
                try:
                    _dbg("Ddlg3w", "wait tp dialog gone (end)", {})
                except Exception:
                    pass
        except Exception:
            pass

        # Guard: không để stuck quá lâu ở bước dismiss
        _dbg("Ddlg9", "dismiss dialogs exit", {"ms": int((time.time() - t0) * 1000)})
    except Exception:
        pass


def _handle_tou_interstitial(driver, log_callback=None, max_wait_s: float = 15.0) -> bool:
    """
    Thoát trang điều khoản/consent của YouTube Studio nếu bị redirect tới /tou/interstitial.
    Trả về True nếu đã xử lý (có can thiệp/click) hoặc đã thoát khỏi interstitial; False nếu không thấy.
    """
    try:
        t0 = time.time()
        acted = False
        while (time.time() - t0) < max_wait_s:
            try:
                url = (driver.current_url or "")
            except Exception:
                url = ""

            try:
                is_tou = bool(
                    driver.execute_script(
                        "try{"
                        "var u=(location&&location.href)?String(location.href):'';"
                        "if(u.includes('/tou/')||u.includes('interstitial')) return true;"
                        "// Một số trang consent không hiện rõ URL, fallback theo button text"
                        "var body=(document.body&&document.body.innerText)?document.body.innerText:'';"
                        "return body.includes('Điều khoản')||body.includes('Terms')||body.includes('consent');"
                        "}catch(e){return false;}"
                    )
                )
            except Exception:
                is_tou = ("/tou/" in url) or ("interstitial" in url)

            if not is_tou:
                if acted:
                    _dbg("TOU1", "tou interstitial resolved", {"url": (url or "")[:160]})
                return acted

            # đang ở interstitial
            if not acted:
                _dbg("TOU0", "tou interstitial detected", {"url": (url or "")[:160]})
                _log(log_callback, "⚠️ YouTube yêu cầu xác nhận điều khoản/consent (tou interstitial) — đang tự xử lý...")
            acted = True

            # JS-click các nút đồng ý/tiếp tục phổ biến
            try:
                clicked = bool(
                    driver.execute_script(
                        "try{"
                        "var texts=["
                        "'Tôi đồng ý','Tôi chấp nhận','Đồng ý','Chấp nhận','Tiếp tục','Tiếp tục','OK','Đóng',"
                        "'I agree','I accept','Accept','Agree','Continue','Next','Got it'"
                        "];"
                        "var btns=[...document.querySelectorAll('button, ytcp-button, tp-yt-paper-button')];"
                        "function norm(s){return (s||'').toString().replace(/\\s+/g,' ').trim().toLowerCase();}"
                        "for(var b of btns){"
                        "  var t=norm(b.getAttribute&&b.getAttribute('aria-label'))||norm(b.innerText);"
                        "  for(var x of texts){"
                        "    var nx=norm(x);"
                        "    if(t && (t===nx || t.includes(nx))){"
                        "      try{b.scrollIntoView({block:'center'});}catch(e){}"
                        "      try{b.click();}catch(e){try{(b.querySelector('button')||b).click();}catch(e2){}}"
                        "      return true;"
                        "    }"
                        "  }"
                        "}"
                        "return false;"
                        "}catch(e){return false;}"
                    )
                )
            except Exception:
                clicked = False

            _dbg("TOU2", "tou interstitial click attempt", {"clicked": bool(clicked), "url": (url or "")[:160]})

            # sau click, chờ redirect một chút; nếu không có click được thì cũng chờ ngắn rồi vòng lại
            time.sleep(1.0)

            # Nếu vẫn ở interstitial, thử load lại studio URL (không dùng /channel/upload theo yêu cầu)
            try:
                if (driver.current_url or "").find("/tou/") != -1:
                    driver.get(YOUTUBE_STUDIO_URL)
            except Exception:
                pass

        # hết budget mà vẫn ở interstitial
        try:
            _dbg("TOU9", "tou interstitial not resolved", {"url": (driver.current_url or "")[:160]})
        except Exception:
            pass
        return acted
    except Exception:
        return False


# Kích thước lô upload (Studio chọn nhiều file một lần).
UPLOAD_BATCH_SIZE = 15

def _wait_upload_form_controls(driver, log_callback=None):
    """Chờ ô tiêu đề / radio kids sau khi đã chọn file."""
    t_controls = time.time()
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "#title-textarea div#textbox[contenteditable='true'], "
                "ytcp-social-suggestions-textbox#title-textarea div#textbox, "
                "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_MFK'], "
                "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_NOT_MFK']"
            ))
        )
    except Exception:
        time.sleep(2)
    _dbg("S3", "post-file controls ready", {"ms": int((time.time() - t_controls) * 1000)})

def _upload_dialog_text_has_filename(dialog_text_lower: str, basename: str) -> bool:
    """
    Studio thường cắt tên dài (…) nên không thể dùng `full basename in innerText`.
    Khớp: full name, stem, đuôi định danh, hoặc đầu+đuôi khi tên rất dài.
    """
    txt = dialog_text_lower or ""
    bn = (os.path.basename(basename) or "").strip().lower()
    if not bn:
        return True
    if bn in txt:
        return True
    stem, _ext = os.path.splitext(bn)
    if not stem:
        return False
    if stem in txt:
        return True
    if len(stem) > 36:
        if stem[:32] in txt and stem[-18:] in txt:
            return True
    for tail_len in (40, 32, 24, 16):
        if len(stem) >= tail_len and stem[-tail_len:] in txt:
            return True
    # Token dài (đuôi hash / id) — tránh khớp chung "LiveScore", "2026" cho cả lô
    for part in reversed(stem.split("_")):
        if len(part) >= 10 and part in txt:
            return True
    return False


def _select_upload_dialog_row_by_basename(driver, basename, log_callback=None):
    """Chọn một video trong dialog upload đa file (theo tên file hoặc stem)."""
    stem, _ext = os.path.splitext(basename or "")
    stem_l = (stem or "").lower()
    base_l = (basename or "").lower()
    try:
        ok = driver.execute_script(
            _JS_YTB_UPLOAD_PROGRESS_ROOT
            + """
            var stem = arguments[0], base = arguments[1];
            var d = __ytbUploadProgressRoot();
            if (!d) return false;
            var candidates = d.querySelectorAll(
              'ytcp-video-upload-progress, ytcp-ve-video-result-row, ytcp-ve-row, '
              + 'tp-yt-paper-item, a, div[role=\"button\"], ytcp-entity-card'
            );
            function norm(s){ return (s||'').toString().toLowerCase(); }
            for (var i = 0; i < candidates.length; i++) {
              var el = candidates[i];
              var t = norm(el.textContent) + ' ' + norm(el.getAttribute && el.getAttribute('title'));
              if (base && t.indexOf(base) !== -1) {
                try { el.scrollIntoView({block:'center'}); el.click(); return true; } catch(e) {}
              }
              if (stem && t.indexOf(stem) !== -1) {
                try { el.scrollIntoView({block:'center'}); el.click(); return true; } catch(e) {}
              }
            }
            return false;
            """,
            stem_l,
            base_l,
        )
        return bool(ok)
    except Exception:
        return False


def _open_multiupload_editor_by_basename(driver, basename, log_callback=None, timeout_s=90):
    """
    Trong dialog upload nhiều file (#progress-list), bấm nút Chỉnh sửa của đúng video.
    Ưu tiên match theo progress-title và aria-label để đi từ trên xuống ổn định.
    """
    bn = os.path.basename(basename or "")
    if not bn:
        return False
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        try:
            ok = driver.execute_script(
                _JS_YTB_UPLOAD_PROGRESS_ROOT
                + """
                var base = (arguments[0] || '').toLowerCase();
                var stem = base;
                var p = stem.lastIndexOf('.');
                if (p > 0) stem = stem.slice(0, p);
                function norm(s){ return (s || '').toString().toLowerCase(); }
                function matchName(name){
                  var t = norm(name);
                  if (!t) return false;
                  if (base && t.indexOf(base) !== -1) return true;
                  if (stem && t.indexOf(stem) !== -1) return true;
                  if (stem && stem.length > 36) {
                    var head = stem.slice(0, 32), tail = stem.slice(-18);
                    if (head && tail && t.indexOf(head) !== -1 && t.indexOf(tail) !== -1) return true;
                  }
                  var parts = stem.split('_');
                  for (var i = parts.length - 1; i >= 0; i--) {
                    if (parts[i] && parts[i].length >= 10 && t.indexOf(parts[i]) !== -1) return true;
                  }
                  return false;
                }
                var dlg = __ytbUploadProgressRoot();
                if (!dlg) return false;
                var rows = dlg.querySelectorAll('#progress-list li.row, li.row');
                for (var i = 0; i < rows.length; i++) {
                  var row = rows[i];
                  var titleEl = row.querySelector('.progress-title');
                  var t = norm(titleEl ? titleEl.textContent : '');
                  var editBtn = row.querySelector('button.edit-button, ytcp-icon-button.edit-icon');
                  var aria = '';
                  if (editBtn) {
                    aria = norm(editBtn.getAttribute && editBtn.getAttribute('aria-label'));
                    if (!aria) {
                      var realBtn = editBtn.matches && editBtn.matches('button') ? editBtn : editBtn.querySelector('button');
                      aria = norm(realBtn && realBtn.getAttribute ? realBtn.getAttribute('aria-label') : '');
                    }
                  }
                  if (matchName(t) || matchName(aria)) {
                    try {
                      if (editBtn && editBtn.scrollIntoView) editBtn.scrollIntoView({block:'center'});
                      if (editBtn && editBtn.click) { editBtn.click(); return true; }
                    } catch(e) {}
                    try {
                      var b = row.querySelector('button.edit-button');
                      if (b) { b.scrollIntoView({block:'center'}); b.click(); return true; }
                    } catch(e2) {}
                  }
                }
                return false;
                """,
                bn,
            )
            if ok:
                return True
        except Exception:
            pass
        time.sleep(0.7)
    _log(log_callback, f"⚠️ Không bấm được nút Chỉnh sửa trong progress-list cho: {bn}")
    return False


def _open_multiupload_editor_by_index(driver, row_index, log_callback=None, timeout_s=90):
    """
    Bấm nút Chỉnh sửa theo thứ tự dòng trong #progress-list (từ trên xuống).
    Dùng khi dialog đã hoàn tất upload để xử lý tuần tự ổn định.
    """
    try:
        idx = int(row_index)
    except Exception:
        return False
    if idx < 0:
        return False
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        try:
            ok = driver.execute_script(
                _JS_YTB_UPLOAD_PROGRESS_ROOT
                + """
                var idx = arguments[0];
                var d = __ytbUploadProgressRoot();
                if(!d) return false;
                var rows = d.querySelectorAll('#progress-list li.row, li.row');
                if(!rows || rows.length <= idx) return false;
                var row = rows[idx];
                var btn = row.querySelector('button.edit-button');
                if(!btn){
                  var iconHost = row.querySelector('ytcp-icon-button.edit-icon');
                  if(iconHost){
                    btn = iconHost.querySelector('button') || iconHost;
                  }
                }
                if(!btn) return false;
                try{ btn.scrollIntoView({block:'center'}); }catch(e){}
                try{ btn.click(); return true; }catch(e1){}
                try{
                  var ev = new MouseEvent('click', {bubbles:true, cancelable:true, view:window});
                  btn.dispatchEvent(ev);
                  return true;
                }catch(e2){}
                return false;
                """,
                idx,
            )
            if ok:
                return True
        except Exception:
            pass
        time.sleep(0.6)
    _log(log_callback, f"⚠️ Không bấm được nút Chỉnh sửa theo index {idx} trong progress-list.")
    return False


def _get_progress_list_row_count(driver) -> int:
    """Đếm số dòng hiện có trong progress-list của multi-upload dialog (0 nếu không thấy)."""
    try:
        n = driver.execute_script(
            _JS_YTB_UPLOAD_PROGRESS_ROOT
            + "try{var d=__ytbUploadProgressRoot(); if(!d) return 0;"
              "var rows=d.querySelectorAll('#progress-list li.row, li.row');"
              "return rows?rows.length:0;}catch(e){return 0;}"
        )
        return int(n or 0)
    except Exception:
        return 0


def _wait_multi_upload_dialog_ready(driver, expected_files, expected_basenames=None, log_callback=None, timeout_s=180):
    """
    Chờ dialog upload ở trạng thái có thể xử lý tiếp.
    Với lô nhiều file: chỉ cần chờ header «Đã hoàn tất quá trình tải lên» (hoặc mọi dòng 100%),
    rồi bấm Chỉnh sửa từng dòng.
    Dialog có thể là ytcp-uploads-dialog hoặc tp-yt-paper-dialog#dialog (ytcp-multi-progress-monitor);
    tìm root qua #progress-list để không bị chờ vô hạn khi UI không dùng ytcp-uploads-dialog.
    """
    def _dialog_completed_state():
        """
        Trả về (is_completed, visible_rows).
        Completed khi header báo "Đã hoàn tất quá trình tải lên"/"uploads complete"
        hoặc tất cả dòng hiện tại đều là 100%/done.
        """
        try:
            return tuple(
                driver.execute_script(
                    _JS_YTB_UPLOAD_PROGRESS_ROOT
                    + """
                    try{
                      var d=__ytbUploadProgressRoot();
                      if(!d) return [false, 0];
                      var hdrEl =
                        d.querySelector('#expand-button .count')
                        || d.querySelector('.header .count')
                        || d.querySelector('.count');
                      var hdr=(hdrEl||{}).textContent||'';
                      var h=(hdr||'').toLowerCase();
                      var completed=(h.indexOf('đã hoàn tất quá trình tải lên')!==-1
                        || h.indexOf('hoan tat qua trinh tai len')!==-1
                        || h.indexOf('uploads complete')!==-1
                        || h.indexOf('all uploads complete')!==-1
                        || h.indexOf('upload complete')!==-1);
                      var rows=d.querySelectorAll('#progress-list li.row, li.row');
                      var n=rows?rows.length:0;
                      if(!completed && n>0){
                        var doneCount=0;
                        for(var i=0;i<rows.length;i++){
                          var r=rows[i];
                          var st=(r.querySelector('.progress-status-text')||{}).textContent||'';
                          var s=(st||'').toLowerCase();
                          var ok=(s.indexOf('100%')!==-1
                            || s.indexOf('đã tải lên')!==-1
                            || s.indexOf('da tai len')!==-1
                            || s.indexOf('uploaded')!==-1
                            || s.indexOf('complete')!==-1);
                          if(ok) doneCount++;
                        }
                        if(doneCount===rows.length) completed=true;
                        if(!completed && doneCount>0 && doneCount===n) completed=true;
                      }
                      if(!completed){
                        var allTxt = (d.innerText||d.textContent||'').toLowerCase();
                        if(allTxt.indexOf('đã hoàn tất quá trình tải lên')!==-1
                          || allTxt.indexOf('hoan tat qua trinh tai len')!==-1
                          || allTxt.indexOf('uploads complete')!==-1){
                          completed = true;
                        }
                      }
                      return [!!completed, n];
                    }catch(e){return [false,0];}
                    """
                )
            )
        except Exception:
            return (False, 0)

    deadline = time.time() + timeout_s
    last_row_log = -1
    while time.time() < deadline:
        completed, visible_rows = _dialog_completed_state()

        if expected_files > 1 and log_callback and visible_rows != last_row_log and visible_rows > 0:
            last_row_log = visible_rows
            _log(
                log_callback,
                f"Đã thấy {visible_rows} dòng trong progress-list — chờ «Đã hoàn tất quá trình tải lên»…",
            )

        if completed:
            if log_callback:
                _log(
                    log_callback,
                    f"Đã hoàn tất tải lên ({visible_rows} dòng trong progress-list) — bắt đầu Chỉnh sửa từng video.",
                )
            return True

        if expected_files <= 1:
            return True

        time.sleep(1.5)

    _log(
        log_callback,
        "⚠️ Dialog chưa báo hoàn tất trong thời gian chờ — vẫn thử xử lý Chỉnh sửa từng dòng.",
    )
    _report_chrome_console_for_debug(
        driver,
        log_callback,
        note="sau timeout chờ dialog upload",
    )
    return False

def _open_studio_and_get_file_input(driver, log_callback=None):
    """Mở Studio, bấm Upload, trả về input[type=file]."""
    wait = WebDriverWait(driver, 60)
    # Mở YouTube Studio
    _log(log_callback, "Đang mở YouTube Studio...")
    # Nếu đang ở sẵn Studio thì không reload để tiết kiệm thời gian
    try:
        if not (driver.current_url or "").startswith(YOUTUBE_STUDIO_URL):
            driver.get(YOUTUBE_STUDIO_URL)
    except Exception:
        driver.get(YOUTUBE_STUDIO_URL)
    # Máy khác có thể bị redirect tới /tou/interstitial -> xử lý ngay để tránh FAIL hàng loạt
    _handle_tou_interstitial(driver, log_callback=log_callback, max_wait_s=15.0)
    # Chờ UI Studio sẵn sàng (upload button hoặc điều hướng upload page)
    t_studio = time.time()
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "ytcp-button#upload-icon, [aria-label*='Upload'], [aria-label*='upload'], tp-yt-paper-button#upload-icon, #upload-icon"
            ))
        )
    except Exception:
        pass
    _dbg("S1", "studio ready waited", {"ms": int((time.time() - t_studio) * 1000), "url": (driver.current_url or "")[:120]})

    # Fast path: nếu đã ở Studio và nút Upload dùng được thì bấm luôn để ra input[type=file]
    # (giảm thời gian chờ/dọn dialog không cần thiết).
    fast_file_input = False
    try:
        t_fast = time.time()
        _dbg("Q0", "fast-path try open upload", {"url": (driver.current_url or "")[:160]})
        driver.execute_script(
            "try{"
            "var btn=document.querySelector('#upload-icon');"
            "if(btn){btn.scrollIntoView({block:'center'}); btn.click(); return true;}"
            "return false;"
            "}catch(e){return false;}"
        )
        try:
            try:
                driver.implicitly_wait(0)
            except Exception:
                pass
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
            _dbg("Q1", "fast-path got file input", {"ms": int((time.time() - t_fast) * 1000)})
            fast_file_input = True
        except Exception:
            _dbg("Q1", "fast-path no file input", {"ms": int((time.time() - t_fast) * 1000)})
        finally:
            try:
                driver.implicitly_wait(10)
            except Exception:
                pass
    except Exception:
        pass

    # Nếu fast-path đã có file input thì đi thẳng tới bước chọn file, không dọn dialog / không điều hướng nữa.
    if not fast_file_input:
        # Nếu đang còn dialog của lần upload trước (uploads/prechecks), ưu tiên dọn + refresh rồi thử fast-path lại.
        try:
            blockers = driver.execute_script(
                "try{return {"
                "uploadsDialog: !!document.querySelector('ytcp-uploads-dialog'),"
                "prechecksDialog: !!document.querySelector('ytcp-prechecks-warning-dialog'),"
                "tpDialog: !!document.querySelector('tp-yt-paper-dialog')"
                "};}catch(e){return {err:String(e).slice(0,120)};}"
            ) or {}
        except Exception:
            blockers = {}
        if blockers.get("uploadsDialog") or blockers.get("prechecksDialog") or blockers.get("tpDialog"):
            try:
                _dbg("Q2", "fast-path blocked; dismiss+refresh", blockers)
            except Exception:
                pass
            _dismiss_blocking_dialogs(driver, log_callback=log_callback)
            try:
                driver.refresh()
            except Exception:
                try:
                    driver.get(YOUTUBE_STUDIO_URL)
                except Exception:
                    pass
            _handle_tou_interstitial(driver, log_callback=log_callback, max_wait_s=10.0)
            # thử lại click upload-icon sau refresh
            try:
                t_retry = time.time()
                driver.execute_script(
                    "try{var btn=document.querySelector('#upload-icon');"
                    "if(btn){btn.scrollIntoView({block:'center'}); btn.click();}}catch(e){}"
                )
                WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
                _dbg("Q3", "retry after refresh got file input", {"ms": int((time.time() - t_retry) * 1000)})
                fast_file_input = True
            except Exception:
                _dbg("Q3", "retry after refresh no file input", {"ms": int((time.time() - t_retry) * 1000)})

    if not fast_file_input:
        # Dọn dialog/overlay còn sót lại từ lần trước để tránh intercept/chờ vô ích
        # JS probe để tránh find_elements bị stall (runtime evidence ~20s).
        try:
            d0 = driver.execute_script(
                "try{return {"
                "uploadsDialog: !!document.querySelector('ytcp-uploads-dialog'),"
                "prechecksDialog: !!document.querySelector('ytcp-prechecks-warning-dialog'),"
                "tpDialog: !!document.querySelector('tp-yt-paper-dialog')"
                "};}catch(e){return {err:String(e).slice(0,120)};}"
            )
            _dbg("D0", "before dismiss dialogs", d0 or {})
        except Exception:
            pass
        _dismiss_blocking_dialogs(driver, log_callback=log_callback)
        try:
            d1 = driver.execute_script(
                "try{return {"
                "uploadsDialog: !!document.querySelector('ytcp-uploads-dialog'),"
                "prechecksDialog: !!document.querySelector('ytcp-prechecks-warning-dialog'),"
                "tpDialog: !!document.querySelector('tp-yt-paper-dialog')"
                "};}catch(e){return {err:String(e).slice(0,120)};}"
            )
            _dbg("D1", "after dismiss dialogs", d1 or {})
        except Exception:
            pass

    if fast_file_input:
        # Đã có input file từ fast-path -> đi thẳng chọn file (không điều hướng / không fallback).
        t_filewait = time.time()
        _dbg("C2", "after goto upload (fast-path)", {"url": (driver.current_url or "")[:200]})
        _dbg("S2", "file input wait (fast-path)", {"ms": int((time.time() - t_filewait) * 1000), "url": (driver.current_url or "")[:120]})
    else:
        # Bỏ hẳn driver.get(/channel/upload) theo yêu cầu.
        # Chiến lược: click Upload -> nếu fail thì dismiss + refresh -> click Upload lại.
        t_filewait = time.time()
        _dbg("C2", "open upload (click-only)", {"url": (driver.current_url or "")[:200]})
        got_input = False
        for attempt in (1, 2):
            try:
                u0 = driver.execute_script(
                    "try{return {"
                    "url:(location&&location.href)?location.href.slice(0,160):'',"
                    "uploadsDialog: !!document.querySelector('ytcp-uploads-dialog'),"
                    "prechecksDialog: !!document.querySelector('ytcp-prechecks-warning-dialog'),"
                    "tpDialog: !!document.querySelector('tp-yt-paper-dialog')"
                    "};}catch(e){return {err:String(e).slice(0,120)};}"
                )
                _dbg("Ux0", f"upload click attempt {attempt}", u0 or {"url": (driver.current_url or "")[:160]})
            except Exception:
                pass

            if attempt == 2:
                _dismiss_blocking_dialogs(driver, log_callback=log_callback)
                try:
                    driver.refresh()
                except Exception:
                    pass

            try:
                upload_btn = wait.until(EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "ytcp-button#upload-icon, [aria-label*='Upload'], [aria-label*='upload'], "
                    "tp-yt-paper-button#upload-icon, #upload-icon"
                )))
                try:
                    probe = driver.execute_script(
                        "var el=document.querySelector('#upload-icon');"
                        "if(!el) return {has:false};"
                        "var r=el.getBoundingClientRect();"
                        "var x=Math.floor(r.left+r.width/2), y=Math.floor(r.top+r.height/2);"
                        "var top=document.elementFromPoint(x,y);"
                        "function s(e){if(!e) return null; return {tag:e.tagName,id:e.id,cls:(e.className||'').toString().slice(0,80),aria:e.getAttribute && (e.getAttribute('aria-label')||'')};}"
                        "return {has:true,rect:{x:x,y:y,w:Math.floor(r.width),h:Math.floor(r.height)},top:s(top),self:s(el)};"
                    )
                    _dbg("Ux1", "upload-icon top element probe", probe or {})
                except Exception:
                    pass
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", upload_btn)
                driver.execute_script("arguments[0].click();", upload_btn)

                try:
                    try:
                        driver.implicitly_wait(0)
                    except Exception:
                        pass
                    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
                    got_input = True
                    _dbg("Ux2", "upload click led to file input", {"attempt": attempt, "ms": int((time.time() - t_filewait) * 1000)})
                    break
                finally:
                    try:
                        driver.implicitly_wait(10)
                    except Exception:
                        pass
            except Exception:
                continue

        if not got_input:
            _dbg("Ux2", "failed to obtain file input", {"ms": int((time.time() - t_filewait) * 1000)})
            raise TimeoutException("Không mở được form upload (không thấy input[type=file]).")

        _dbg("S2", "file input wait", {"ms": int((time.time() - t_filewait) * 1000), "url": (driver.current_url or "")[:120]})

    # Chọn file: input[type="file"]
    file_input = wait.until(EC.presence_of_element_located((
        By.CSS_SELECTOR, "input[type='file']"
    )))
    return file_input

def _fill_metadata_to_done(
    driver,
    file_path_abs,
    video_title,
    made_for_kids,
    visibility,
    log_callback,
    on_link_available,
):
    """Sau khi file đã chọn và form đã sẵn sàng: tiêu đề → trẻ em → … → Lưu."""
    result = {"success": False, "url": None, "error": None, "excel_done": False}
    wait = WebDriverWait(driver, 60)
    done_save_clicked = False
    try:

        # Tiêu đề: có nhập trong cấu hình thì điền; không nhập thì bỏ qua, chạy luôn bước Có/Không trẻ em (radio đã có sẵn)
        # YouTube Studio dùng div contenteditable cho ô tiêu đề, không phải input
        if video_title and str(video_title).strip():
            try:
                title_selector = (
                    "#title-textarea div#textbox[contenteditable='true'], "
                    "ytcp-social-suggestions-textbox#title-textarea div#textbox, "
                    "div#textbox[role='textbox'][aria-label*='tiêu đề'], "
                    "[aria-label*='Thêm tiêu đề']"
                )
                title_el = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, title_selector))
                )
                title_el.click()
                time.sleep(0.3)
                # Clear contenteditable: Ctrl+A rồi gõ mới
                title_el.send_keys(Keys.CONTROL + "a")
                title_el.send_keys(Keys.BACKSPACE)
                time.sleep(0.2)
                title_el.send_keys(str(video_title).strip())
                _log(log_callback, "Đã điền tiêu đề.")
            except TimeoutException:
                _log(log_callback, "Không tìm thấy ô tiêu đề, bỏ qua (có thể dùng tên file).")
        else:
            _log(log_callback, "Không điền tiêu đề (để trống), chạy luôn phần Có/Không trẻ em.")

        # Bước 2: Có / Không - Nội dung dành cho trẻ em
        want_yes = bool(made_for_kids)
        try:
            # #region agent log
            try:
                _agent_debug_log(
                    "K1",
                    "Selecting made_for_kids radio",
                    {"madeForKids": want_yes, "url": driver.current_url},
                    run_id="kids_select_pre",
                )
            except Exception:
                pass
            # #endregion agent log
            # Chờ cả hai radio có trên trang rồi mới chọn (tránh click nhầm)
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_MFK']"
            )))
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_NOT_MFK']"
            )))
            time.sleep(0.5)
            yes_el = driver.find_element(By.CSS_SELECTOR, "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_MFK']")
            no_el = driver.find_element(By.CSS_SELECTOR, "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_NOT_MFK']")
            target = yes_el if want_yes else no_el
            # #region agent log
            try:
                _agent_debug_log(
                    "K1b",
                    "Kids radio about to click",
                    {"target": "MFK" if want_yes else "NOT_MFK", "want_yes": want_yes},
                    run_id="kids_click_pre",
                )
            except Exception:
                pass
            # #endregion agent log
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target)
            time.sleep(0.3)
            WebDriverWait(driver, 8).until(EC.element_to_be_clickable(target))
            try:
                target.click()
            except Exception:
                driver.execute_script("arguments[0].click();", target)
            time.sleep(0.8)
            yes_checked = (yes_el.get_attribute("aria-checked") or "").strip().lower()
            no_checked = (no_el.get_attribute("aria-checked") or "").strip().lower()
            try:
                _agent_debug_log(
                    "K2",
                    "Kids radio aria-checked after click",
                    {"yes": yes_checked, "no": no_checked, "expected": "yes" if want_yes else "no"},
                    run_id="kids_select_post",
                )
            except Exception:
                pass
            if want_yes and yes_checked != "true":
                try:
                    driver.execute_script("arguments[0].click();", yes_el)
                    time.sleep(0.5)
                except Exception:
                    pass
            if (not want_yes) and no_checked != "true":
                try:
                    driver.execute_script("arguments[0].click();", no_el)
                    time.sleep(0.5)
                except Exception:
                    pass
            # #region agent log — trạng thái cuối sau retry
            try:
                yes_checked_final = (yes_el.get_attribute("aria-checked") or "").strip().lower()
                no_checked_final = (no_el.get_attribute("aria-checked") or "").strip().lower()
                _agent_debug_log(
                    "K3",
                    "Kids radio final state after retry",
                    {"yes": yes_checked_final, "no": no_checked_final, "want_yes": want_yes},
                    run_id="kids_select_final",
                )
            except Exception:
                pass
            # #endregion agent log
            if want_yes:
                _log(log_callback, "Đã chọn: Có, nội dung dành cho trẻ em.")
            else:
                _log(log_callback, "Đã chọn: Không, nội dung không dành cho trẻ em.")
        except Exception:
            try:
                if want_yes:
                    kids_yes = driver.find_elements(By.CSS_SELECTOR, "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_MFK']")
                    if not kids_yes:
                        kids_yes = driver.find_elements(By.XPATH, "//*[contains(@aria-label,'trẻ em') or contains(@aria-label,'kids') or contains(text(),'Có') or contains(text(),'Yes')]")
                    for el in kids_yes:
                        try:
                            if "MFK" in (el.get_attribute("name") or "") or "kids" in (el.get_attribute("aria-label") or "").lower():
                                el.click()
                                _log(log_callback, "Đã chọn: Có, nội dung dành cho trẻ em.")
                                break
                        except Exception:
                            pass
                else:
                    kids_no = driver.find_elements(By.CSS_SELECTOR, "tp-yt-paper-radio-button[name='VIDEO_MADE_FOR_KIDS_NOT_MFK']")
                    if not kids_no:
                        kids_no = driver.find_elements(By.XPATH, "//*[contains(@aria-label,'Không') or contains(@aria-label,'No') or contains(text(),'Không') or contains(text(),'No')]")
                    for el in kids_no:
                        try:
                            if "NOT_MFK" in (el.get_attribute("name") or "") or "không" in (el.get_attribute("aria-label") or "").lower() or "no" in (el.get_attribute("aria-label") or "").lower():
                                el.click()
                                _log(log_callback, "Đã chọn: Không, nội dung không dành cho trẻ em.")
                                break
                        except Exception:
                            pass
            except Exception:
                pass
        time.sleep(1)

        # YouTube Studio thường bật nút Next vài giây sau khi chọn Có/Không trẻ em — chờ để tránh bấm quá sớm
        _wait_studio_next_enabled(driver, timeout_s=60, log_callback=log_callback)
        time.sleep(0.5)

        # Bước 3–4: Next hai lần — lần 1: Chi tiết → Các thành phần của video; lần 2: sang Kiểm tra ban đầu
        if not _click_next(driver, log_callback, "Chi tiết → Các thành phần của video"):
            result["error"] = "Không chuyển được sang bước Các thành phần của video"
            _log(log_callback, f"❌ {result['error']}. Chuyển video khác.")
            return result
        time.sleep(2)
        if not _click_next(driver, log_callback, "Các thành phần của video → Kiểm tra ban đầu"):
            result["error"] = "Không chuyển được sang bước Kiểm tra ban đầu"
            _log(log_callback, f"❌ {result['error']}. Chuyển video khác.")
            return result
        time.sleep(2)

        # Bước 5: Kiểm tra ban đầu — không báo gì thì Next; có bản quyền thì xử lý thay thế bài hát
        if not _handle_checks_and_copyright(driver, wait, log_callback):
            result["error"] = "Không hoàn thành bước kiểm tra/bản quyền"
            return result
        time.sleep(1)

        # Một số luồng (đặc biệt khi xử lý bản quyền / thay thế bài hát) có thể đã đưa thẳng tới bước
        # Chế độ hiển thị, nên không còn nút Next nữa. Nếu đã thấy #privacy-radios thì bỏ qua click Next.
        already_on_visibility = False
        try:
            try:
                _agent_debug_log(
                    "V1",
                    "Checking if already on visibility step (#privacy-radios)",
                    {"url": driver.current_url},
                    run_id="visibility_check_pre",
                )
            except Exception:
                pass
            # Tránh implicit_wait=10s làm check này bị kéo dài
            try:
                driver.implicitly_wait(0)
            except Exception:
                pass
            try:
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#privacy-radios"))
                )
            finally:
                try:
                    driver.implicitly_wait(10)
                except Exception:
                    pass
            # #region agent log
            try:
                el = driver.find_element(By.CSS_SELECTOR, "#privacy-radios")
                uploads_dialog_present = bool(driver.find_elements(By.CSS_SELECTOR, "ytcp-uploads-dialog"))
                _agent_debug_log(
                    "V1",
                    "Found #privacy-radios during visibility check",
                    {
                        "url": driver.current_url,
                        "privacyRadiosDisplayed": bool(getattr(el, "is_displayed", lambda: None)()),
                        "uploadsDialogPresent": uploads_dialog_present,
                    },
                    run_id="visibility_check_detail",
                )
            except Exception:
                pass
            # #endregion agent log
            already_on_visibility = True
            try:
                _agent_debug_log(
                    "V1",
                    "Detected visibility step already present; skipping Next",
                    {"url": driver.current_url},
                    run_id="visibility_check_post",
                )
            except Exception:
                pass
        except Exception:
            already_on_visibility = False

        if not already_on_visibility:
            if not _click_next(driver, log_callback, "Sau kiểm tra ban đầu"):
                result["error"] = "Không chuyển được sang Chế độ hiển thị"
                _log(log_callback, f"❌ {result['error']}. Chuyển video khác.")
                return result
            time.sleep(2)

        # Bước 6: Chế độ hiển thị (ytcp-uploads-review) — chọn Riêng tư / Không công khai / Công khai, rồi Lưu và lấy link
        # Radio: tp-yt-paper-radio-button name="PRIVATE"|"UNLISTED"|"PUBLIC" trong #privacy-radios
        visibility_name_map = {
            "public": "PUBLIC",
            "unlisted": "UNLISTED",
            "private": "PRIVATE",
        }
        visibility_label_map = {
            "public": ["Công khai", "Public"],
            "unlisted": ["Không công khai", "Unlisted"],
            "private": ["Riêng tư", "Private"],
        }
        visibility_key = visibility.lower()
        visibility_name = visibility_name_map.get(visibility_key, "UNLISTED")
        visibility_labels = visibility_label_map.get(visibility_key, visibility_label_map["unlisted"])

        chosen = False
        # Ưu tiên chọn trực tiếp radio trong nhóm #privacy-radios theo name
        try:
            # #region agent log
            try:
                _agent_debug_log(
                    "V2",
                    "Attempting visibility select by name",
                    {"visibilityKey": visibility_key, "visibilityName": visibility_name, "url": driver.current_url},
                    run_id="visibility_select_pre",
                )
            except Exception:
                pass
            # #endregion agent log
            radio = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    f"#privacy-radios tp-yt-paper-radio-button[name='{visibility_name}']"
                ))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", radio)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
                By.CSS_SELECTOR,
                f"#privacy-radios tp-yt-paper-radio-button[name='{visibility_name}']"
            )))
            try:
                radio.click()
            except Exception:
                driver.execute_script("arguments[0].click();", radio)
            # #region agent log
            try:
                _agent_debug_log(
                    "V2",
                    "Visibility select by name clicked",
                    {
                        "visibilityName": visibility_name,
                        "ariaChecked": radio.get_attribute("aria-checked"),
                        "url": driver.current_url,
                    },
                    run_id="visibility_select_post",
                )
            except Exception:
                pass
            # #endregion agent log
            _log(log_callback, f"Đã chọn visibility (theo name): {visibility_key}")
            chosen = True
        except Exception:
            chosen = False

        # Fallback 1: tìm theo #radioLabel text (tiếng Việt + tiếng Anh)
        if not chosen:
            try:
                radios = driver.find_elements(
                    By.CSS_SELECTOR,
                    "#privacy-radios tp-yt-paper-radio-button"
                )
                for r in radios:
                    try:
                        label_el = r.find_element(By.CSS_SELECTOR, "#radioLabel")
                        label_text = (label_el.text or "").strip()
                        if any(label_text == l for l in visibility_labels):
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", r)
                            WebDriverWait(driver, 5).until(EC.element_to_be_clickable(r))
                            try:
                                r.click()
                            except Exception:
                                driver.execute_script("arguments[0].click();", r)
                            # #region agent log
                            try:
                                _agent_debug_log(
                                    "V3",
                                    "Visibility select by #radioLabel clicked",
                                    {
                                        "labelText": label_text,
                                        "ariaChecked": r.get_attribute("aria-checked"),
                                        "url": driver.current_url,
                                    },
                                    run_id="visibility_select_post",
                                )
                            except Exception:
                                pass
                            # #endregion agent log
                            _log(log_callback, f"Đã chọn visibility (theo #radioLabel): {label_text}")
                            chosen = True
                            break
                    except Exception:
                        continue
            except Exception:
                pass

        # Fallback 2: match theo aria-label/text tổng nếu vẫn chưa chọn được
        if not chosen:
            visibility_map = {
                "public": ["Public", "Công khai", "PUBLIC"],
                "unlisted": ["Unlisted", "Không công khai", "UNLISTED"],
                "private": ["Private", "Riêng tư", "PRIVATE"],
            }
            target_visibility = visibility_map.get(visibility_key, visibility_map["unlisted"])
            for r in driver.find_elements(By.CSS_SELECTOR, "#privacy-radios tp-yt-paper-radio-button, tp-yt-paper-radio-button, [role='radio']"):
                try:
                    label = (r.get_attribute("aria-label") or r.text or "").strip()
                    if any(v in label for v in target_visibility):
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", r)
                        try:
                            r.click()
                        except Exception:
                            driver.execute_script("arguments[0].click();", r)
                        _log(log_callback, f"Đã chọn visibility (theo label fallback): {visibility_key}")
                        chosen = True
                        break
                except Exception:
                    continue

        if not chosen:
            _log(log_callback, "⚠️ Không chọn được radio chế độ hiển thị, YouTube có thể đã đổi giao diện.")
        time.sleep(1)

        # Thử lấy link ngay sau khi chọn visibility (trước khi bấm Lưu) — để log + Excel; chưa coi là upload xong.
        t_early = time.time()
        early_url = _try_get_video_link_from_page(driver)
        _dbg("L1", "early link probe", {"ms": int((time.time() - t_early) * 1000), "hasUrl": bool(early_url), "url": (early_url or "")[:80]})
        if early_url:
            result["url"] = early_url
            _log(log_callback, f"Link video (trước Lưu): {result['url']}")
            if on_link_available:
                try:
                    on_link_available(early_url)
                except Exception:
                    pass
                result["excel_done"] = True

        # Bấm "Lưu" / Done (bước Chế độ hiển thị: #done-button hiện, #next-button ẩn)
        try:
            # Tìm đúng ytcp-button#done-button trong footer như DOM bạn gửi
            done_outer = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "ytcp-uploads-dialog ytcp-button#done-button, ytcp-button#done-button"
                ))
            )
            done_inner = done_outer.find_element(
                By.CSS_SELECTOR,
                "button[aria-label='Lưu'], button.ytcpButtonShapeImplHost"
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", done_inner)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(done_inner))
            try:
                done_inner.click()
            except Exception:
                driver.execute_script("arguments[0].click();", done_inner)
            done_save_clicked = True
            _log(log_callback, "Đã bấm Lưu / Xuất bản ở bước Chế độ hiển thị.")
        except TimeoutException:
            result["success"] = False
            result["error"] = "Không tìm thấy nút Lưu/Done"
            _log(log_callback, f"❌ {result['error']}")
            return result
        except Exception as e:
            result["success"] = False
            result["error"] = f"Lỗi nút Lưu/Done: {e}"
            _log(log_callback, f"❌ {result['error']}")
            return result

        # Sau khi bấm Done, đôi khi YouTube hiện cảnh báo pre-checks -> cần xác nhận "Vẫn xuất bản"
        try:
            _handle_prechecks_warning_after_done(driver, log_callback=log_callback)
        except Exception:
            pass

        # Sau Done, thay vì sleep cứng: thử lấy link vài lần (vì YouTube có thể render link nhanh/chậm tuỳ video)
        t_poll = time.time()
        for _ in range(10):
            if result.get("url"):
                break
            try:
                url_probe = _try_get_video_link_from_page(driver)
                if url_probe:
                    result["url"] = url_probe
                    _log(log_callback, f"Link video: {result['url']}")
                    if on_link_available and not result.get("excel_done"):
                        try:
                            on_link_available(url_probe)
                        except Exception:
                            pass
                        result["excel_done"] = True
                    break
            except Exception:
                pass
            time.sleep(1)
        _dbg("L2", "post-done link poll", {"ms": int((time.time() - t_poll) * 1000), "hasUrl": bool(result.get("url"))})

        # Lấy link video nếu chưa có (sau khi bấm Done)
        if not result.get("url"):
            try:
                url = _try_get_video_link_from_page(driver)
                if url:
                    result["url"] = url
                    _log(log_callback, f"Link video: {result['url']}")
                    if on_link_available and not result.get("excel_done"):
                        try:
                            on_link_available(url)
                        except Exception:
                            pass
                        result["excel_done"] = True
                else:
                    result["url"] = None
                    _log(log_callback, "Upload có thể đã xong nhưng chưa lấy được link (kiểm tra Studio).")
            except Exception as e:
                result["url"] = None
                result["error"] = str(e)

        if done_save_clicked:
            result["success"] = True
            try:
                _dismiss_post_publish_share_dialog(driver, log_callback=log_callback)
            except Exception:
                pass

    except Exception as e:
        result["success"] = False
        result["error"] = str(e)
        _log(log_callback, f"Lỗi upload: {e}")
        _dbg("S9", "_fill_metadata_to_done exception", {"err": str(e)[:200]})
    return result


def _subprocess_no_window_kwargs():
    if sys.platform == "win32":
        return {"creationflags": getattr(subprocess, "CREATE_NO_WINDOW", 0)}
    return {}


def _probe_video_dimensions(file_path):
    """
    Đọc width x height luồng video đầu tiên (ffprobe). Trả về (w, h) hoặc None.
    """
    ffprobe = shutil.which("ffprobe")
    if not ffprobe or not file_path or not os.path.isfile(file_path):
        return None
    try:
        out = subprocess.run(
            [
                ffprobe,
                "-v",
                "error",
                "-select_streams",
                "v:0",
                "-show_entries",
                "stream=width,height",
                "-of",
                "csv=p=0:s=x",
                os.path.abspath(file_path),
            ],
            capture_output=True,
            text=True,
            timeout=60,
            **_subprocess_no_window_kwargs(),
        )
        line = (out.stdout or "").strip().splitlines()
        if not line:
            return None
        parts = line[0].lower().split("x")
        if len(parts) != 2:
            return None
        w, h = int(parts[0]), int(parts[1])
        if w <= 0 or h <= 0:
            return None
        return (w, h)
    except Exception:
        return None


def _classify_studio_list_kind(file_path):
    """
    Tab danh sách Studio: Video (ngang, vd 16:9) vs Shorts (dọc 9:16 hoặc vuông 1:1).
    """
    dim = _probe_video_dimensions(file_path)
    if not dim:
        return "video"
    w, h = dim
    r = w / float(h)
    if abs(r - 1.0) <= 0.10:
        return "shorts"
    if r < 1.0:
        return "shorts"
    return "video"


def _close_multifile_upload_dialog(driver, log_callback=None):
    """Đóng ytcp-uploads-dialog sau khi đã gửi file (upload vẫn chạy nền)."""
    try:
        clicked = driver.execute_script(
            "try{"
            "var d=document.querySelector('ytcp-uploads-dialog');"
            "if(!d) return false;"
            "var c=d.querySelector("
            "'#close-button button, ytcp-icon-button#close-button button, "
            "button[aria-label*=\"Đóng\"], button[aria-label*=\"Close\"]');"
            "if(c){c.click(); return true;}"
            "return false;"
            "}catch(e){return false;}"
        )
        if clicked:
            _log(log_callback, "Đã đóng hộp thoại upload.")
        time.sleep(1.2)
        return bool(clicked)
    except Exception:
        return False


def _go_to_studio_channel_videos(driver, log_callback=None):
    """Mở trang Nội dung có tab Video / Shorts (tp-yt-paper-tab#video-list-*)."""
    _log(log_callback, "Đang mở danh sách Nội dung (Video / Shorts)...")
    try:
        driver.get(YOUTUBE_STUDIO_URL)
    except Exception:
        pass
    time.sleep(1.2)
    _handle_tou_interstitial(driver, log_callback=log_callback, max_wait_s=10.0)
    try:
        _log(log_callback, "Đang điều hướng tới trang /videos (nếu lấy được channel/c slug)...")
        navigated = driver.execute_script(
            "try{"
            "var href=(location&&location.href)?location.href:'';"
            "var m=href.match(/\\/channel\\/(UC[a-zA-Z0-9_-]+)/);"
            "if(m){"
            "var base='https://studio.youtube.com/channel/'+m[1]+'/videos';"
            "if(!href.includes('/videos')||href.indexOf('/videos/upload')>=0){"
            "location.href=base; return 'goto';"
            "}"
            "if(document.querySelector('tp-yt-paper-tab#video-list-uploads-tab')) return 'ok';"
            "}"
            "var mc=href.match(/studio\\.youtube\\.com\\/c\\/([^\\/?#]+)/);"
            "if(mc){"
            "var basec='https://studio.youtube.com/c/'+mc[1]+'/videos';"
            "if(!href.includes('/videos')||href.indexOf('/videos/upload')>=0){"
            "location.href=basec; return 'goto';"
            "}"
            "if(document.querySelector('tp-yt-paper-tab#video-list-uploads-tab')) return 'ok';"
            "}"
            "var links=document.querySelectorAll('a[href*=\"/videos\"]');"
            "for(var i=0;i<links.length;i++){"
            "var h=(links[i].getAttribute('href')||'');"
            "if(h.indexOf('shorts')>=0&&h.indexOf('/videos')<0) continue;"
            "if(h.indexOf('/videos')<0) continue;"
            "links[i].click(); return 'click';"
            "}"
            "return 'fail';"
            "}catch(e){return 'err';}"
        )
        try:
            _log(log_callback, f"Kết quả điều hướng /videos: {navigated}")
        except Exception:
            pass
        if navigated == "goto":
            time.sleep(3.0)
        elif navigated == "click":
            time.sleep(3.0)
    except Exception:
        pass

    # Fallback mạnh: bấm menu drawer "Nội dung" (ytcp-navigation-drawer)
    try:
        _log(log_callback, "Fallback: thử bấm menu drawer “Nội dung/Content”...")
        clicked_drawer = driver.execute_script(
            "try{"
            "function norm(s){return (s||'').toString().replace(/\\s+/g,' ').trim().toLowerCase();}"
            "var items=[...document.querySelectorAll('ytcp-navigation-drawer tp-yt-paper-icon-item')];"
            "for(var i=0;i<items.length;i++){"
            "  var it=items[i];"
            "  var t=norm(it.innerText||it.textContent);"
            "  if(t.includes('nội dung')||t.includes('content')){"
            "    it.click(); return true;"
            "  }"
            "}"
            "return false;"
            "}catch(e){return false;}"
        )
        if clicked_drawer:
            _log(log_callback, "Đã bấm menu Nội dung.")
            time.sleep(3.0)
        else:
            _log(log_callback, "Không bấm được menu Nội dung (có thể drawer chưa render).")
    except Exception:
        pass
    try:
        _log(log_callback, "Đang chờ tab Video/Shorts xuất hiện...")
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "tp-yt-paper-tab#video-list-uploads-tab, #video-list-uploads-tab")
            )
        )
        _log(log_callback, "Đã thấy tab Video/Shorts.")
    except TimeoutException:
        _log(log_callback, "⚠️ Không thấy tab Video — có thể cần mở tay mục Nội dung.")


def _click_studio_content_tab(driver, kind, log_callback=None):
    """kind: 'video' | 'shorts' — bấm tab Video hoặc Shorts."""
    tab_id = "video-list-shorts-tab" if kind == "shorts" else "video-list-uploads-tab"
    label = "Shorts" if kind == "shorts" else "Video"
    try:
        _log(log_callback, f"Đang bấm tab {label} ({tab_id})...")
        tab = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, f"tp-yt-paper-tab#{tab_id}, #{tab_id}"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tab)
        time.sleep(0.2)
        try:
            tab.click()
        except Exception:
            driver.execute_script("arguments[0].click();", tab)
        _log(log_callback, f"Đã chọn tab {label}.")
        time.sleep(1.0)
        return True
    except Exception:
        _log(log_callback, f"⚠️ Không bấm được tab {label}.")
        return False


def _open_content_video_details(driver, title_fragment, log_callback=None, timeout_s=150):
    """
    Tìm hàng ytcp-video-list-cell-video khớp tiêu đề, hover, bấm #video-details (Chi tiết).
    """
    frag = (title_fragment or "").strip().lower()
    if not frag:
        return False
    _log(log_callback, f"Đang tìm video theo title: «{title_fragment}» để bấm Chi tiết...")
    deadline = time.time() + timeout_s
    last_probe = 0
    while time.time() < deadline:
        cells = driver.find_elements(By.CSS_SELECTOR, "ytcp-video-list-cell-video")
        if log_callback and (time.time() - last_probe) > 10:
            last_probe = time.time()
            _log(log_callback, f"Đang quét danh sách: thấy {len(cells)} dòng (cell) trên trang...")
        for cell in cells:
            try:
                title_el = cell.find_element(By.CSS_SELECTOR, "a#video-title")
                label = (title_el.text or title_el.get_attribute("aria-label") or "").strip().lower()
                if not label:
                    continue
                if frag not in label and label not in frag:
                    continue
                _log(log_callback, f"Tìm thấy dòng khớp: «{(title_el.text or title_el.get_attribute('aria-label') or '').strip()}» — đang mở Chi tiết...")
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cell)
                time.sleep(0.35)
                try:
                    ActionChains(driver).move_to_element(cell).pause(0.25).perform()
                except Exception:
                    pass
                time.sleep(0.25)
                try:
                    btn = cell.find_element(By.CSS_SELECTOR, "ytcp-icon-button#video-details")
                    driver.execute_script("arguments[0].click();", btn)
                except Exception:
                    alt = cell.find_elements(
                        By.CSS_SELECTOR,
                        "ytcp-icon-button[aria-label*='Chi tiết'], ytcp-icon-button[aria-label*='Details']",
                    )
                    if not alt:
                        continue
                    driver.execute_script("arguments[0].click();", alt[0])
                _log(log_callback, "Đã mở Chi tiết từ danh sách.")
                return True
            except Exception:
                continue
        time.sleep(2.5)
    _log(log_callback, f"⚠️ Không tìm thấy «{title_fragment}» trong danh sách trong {timeout_s}s.")
    return False


# JS: dialog «Đã đăng video» đôi khi là ytcp-video-share-dialog, đôi khi chỉ là tp-yt-paper-dialog#dialog (ytcp-dialog).
_JS_YTB_POST_PUBLISH_SHARE = (
    "function __ytbFindPostPublishShareDialog(){"
    "var el=document.querySelector('ytcp-video-share-dialog');"
    "if(el&&el.offsetParent!==null)return el;"
    "var paper=document.querySelector('tp-yt-paper-dialog#dialog');"
    "if(paper&&paper.offsetParent!==null){"
    "var title=paper.querySelector('#dialog-title');"
    "var tx=((title&&title.textContent)||'').toLowerCase().replace(/\\s+/g,' ').trim();"
    "if(tx.indexOf('đã đăng')!==-1||tx.indexOf('da dang')!==-1||"
    "tx.indexOf('video published')!==-1||tx.indexOf('uploaded')!==-1){return paper;}"
    "if(paper.querySelector('ytcp-icon-button#close-icon-button')&&paper.querySelector('#share-url'))return paper;"
    "}"
    "var all=document.querySelectorAll('tp-yt-paper-dialog.ytcp-dialog');"
    "for(var i=0;i<all.length;i++){"
    "var p=all[i];"
    "var ti=p.querySelector('#dialog-title');"
    "var t2=((ti&&ti.textContent)||'').toLowerCase();"
    "if(t2.indexOf('đã đăng')!==-1||t2.indexOf('video published')!==-1)return p;"
    "}"
    "return null;"
    "}"
    "function __ytbDismissPostPublishShareClick(dlg){"
    "if(!dlg)return false;"
    "var host=dlg.querySelector('ytcp-icon-button#close-icon-button');"
    "if(host){"
    "var b=host.querySelector('button');if(!b)b=host;"
    "try{b.scrollIntoView({block:'center'});}catch(z){}"
    "try{b.click();return true;}catch(e1){}"
    "try{var ev=new MouseEvent('click',{bubbles:true,cancelable:true,view:window});b.dispatchEvent(ev);return true;}catch(e2){}"
    "}"
    "var foot=dlg.querySelector('ytcp-button#close-button button, ytcp-button#close-button .ytcpButtonShapeImplHost');"
    "if(foot){try{foot.scrollIntoView({block:'center'});}catch(z){}"
    "try{foot.click();return true;}catch(e3){}}"
    "return false;"
    "}"
)


def _dismiss_post_publish_share_dialog(driver, log_callback=None, timeout_s=15):
    """
    Sau khi xuất bản, Studio mở dialog «Đã đăng video» (chia sẻ).
    Đóng bằng X header (#close-icon-button) hoặc nút «Đóng» footer (#close-button).
    Host có thể là ytcp-video-share-dialog hoặc tp-yt-paper-dialog#dialog — cả hai đều tìm được.
    """
    t0 = time.time()
    deadline = t0 + float(timeout_s)
    saw_dialog = False
    while time.time() < deadline:
        try:
            has_dlg = driver.execute_script(
                _JS_YTB_POST_PUBLISH_SHARE
                + "try{var d=__ytbFindPostPublishShareDialog();"
                "if(!d)return false;"
                "var r=d.getBoundingClientRect();"
                "return d.offsetParent!==null&&r.width>4&&r.height>4;"
                "}catch(e){return false;}"
            )
            if has_dlg:
                saw_dialog = True
            elif not saw_dialog and (time.time() - t0) > 3.0:
                return False
            if not has_dlg:
                time.sleep(0.35)
                continue
            clicked = driver.execute_script(
                _JS_YTB_POST_PUBLISH_SHARE
                + "try{var dlg=__ytbFindPostPublishShareDialog();"
                "return __ytbDismissPostPublishShareClick(dlg);"
                "}catch(e){return false;}"
            )
            if clicked:
                _log(log_callback, "Đã đóng hộp thoại «Đã đăng video» (X hoặc Đóng).")
                time.sleep(0.75)
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


def _close_studio_editor_overlay(driver, log_callback=None):
    """Đóng panel / dialog sau khi Lưu (nút X, Đóng, hoặc Escape)."""
    for _ in range(5):
        try:
            done = driver.execute_script(
                "try{"
                "var sels=["
                "'tp-yt-paper-dialog#dialog ytcp-icon-button#close-icon-button button',"
                "'tp-yt-paper-dialog#dialog ytcp-button#close-button button',"
                "'ytcp-video-share-dialog ytcp-icon-button#close-icon-button button',"
                "'ytcp-video-share-dialog ytcp-button#close-button button',"
                "'ytcp-video-share-dialog #close-icon-button button',"
                "'ytcp-uploads-dialog #close-button button',"
                "'ytcp-uploads-dialog ytcp-icon-button#close-button button',"
                "'ytcp-video-metadata-editor-side-panel #close-button button',"
                "'ytcp-entity-metadata-editor #close-button button',"
                "'tp-yt-paper-dialog button[aria-label*=\"Đóng\"]',"
                "'tp-yt-paper-dialog button[aria-label*=\"Close\"]'"
                "];"
                "for(var i=0;i<sels.length;i++){"
                "var b=document.querySelector(sels[i]);"
                "if(b&&b.offsetParent!==null){b.click(); return true;}"
                "}"
                "var ib=document.querySelector('ytcp-icon-button#close-button button');"
                "if(ib&&ib.offsetParent!==null){ib.click(); return true;}"
                "return false;"
                "}catch(e){return false;}"
            )
            if done:
                _log(log_callback, "Đã đóng panel chỉnh sửa (X).")
                time.sleep(0.9)
                return True
        except Exception:
            pass
        time.sleep(0.35)
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(0.5)
    except Exception:
        pass
    return False


def upload_videos_batch(
    driver,
    file_paths,
    video_title="",
    made_for_kids=False,
    visibility="unlisted",
    log_callback=None,
    on_link_per_file=None,
):
    """
    Upload nhiều file một lần (newline trên Windows), đóng dialog, vào Nội dung (tab Video/Shorts theo tỉ lệ),
    mở Chi tiết từng dòng rồi chạy metadata như luồng đơn; sau mỗi video đóng panel (X).
    on_link_per_file: callable(local_path, url) — gọi khi có link cho từng file.
    Trả về list dict cùng thứ tự file_paths.
    """
    out = []
    if not driver or not file_paths:
        return out
    paths_abs = []
    for p in file_paths:
        ap = os.path.abspath(p)
        if not os.path.isfile(ap):
            out.append({"success": False, "url": None, "error": f"Không có file: {p}", "excel_done": False})
            return out
        paths_abs.append(ap)

    _dbg(
        "B0",
        "upload_videos_batch enter",
        {"n": len(paths_abs), "first": os.path.basename(paths_abs[0])},
    )
    if len(paths_abs) > 1 and video_title and str(video_title).strip():
        _log(
            log_callback,
            "⚠️ Tiêu đề chung cho cả lô: tìm dòng theo cùng một chuỗi — nên để trống tiêu đề để khớp theo tên file.",
        )
    try:
        pre_rows = _get_progress_list_row_count(driver)
        file_input = _open_studio_and_get_file_input(driver, log_callback)
        keys = "\n".join(paths_abs)
        _log(log_callback, "Đang gửi danh sách file vào input[type=file]...")
        file_input.send_keys(keys)
        _log(
            log_callback,
            f"Đã chọn {len(paths_abs)} file video một lần, đang chờ dialog...",
        )
        _wait_multi_upload_dialog_ready(
            driver,
            len(paths_abs),
            expected_basenames=paths_abs,
            log_callback=log_callback,
        )
        if _browser_console_debug_verbose():
            _report_chrome_console_for_debug(
                driver,
                log_callback,
                note="sau khi dialog upload sẵn sàng (YTB_BROWSER_CONSOLE=1)",
            )
        post_rows = _get_progress_list_row_count(driver)
        # Dialog có thể giữ lại video cũ trong cùng phiên => cần offset đúng cho lô mới.
        if post_rows >= (pre_rows + len(paths_abs)):
            batch_start_idx = pre_rows
        else:
            batch_start_idx = max(0, post_rows - len(paths_abs))
        if log_callback:
            _log(
                log_callback,
                f"progress-list hiện có {post_rows} dòng, lô mới bắt đầu từ index {batch_start_idx}.",
            )
        time.sleep(2)
        _log(log_callback, "Đang xử lý ngay trong hộp thoại progress-list (bấm Chỉnh sửa từng video)...")

        for idx, path in enumerate(paths_abs):
            bn = os.path.basename(path)
            if video_title and str(video_title).strip():
                eff_title = str(video_title).strip()
            else:
                eff_title = os.path.splitext(bn)[0]

            _log(
                log_callback,
                f"Lô: video {idx + 1}/{len(paths_abs)} — {bn} (mở từ progress-list)",
            )
            # Ưu tiên basename để tránh click nhầm khi progress-list còn video cũ.
            opened = _open_multiupload_editor_by_basename(driver, bn, log_callback=log_callback, timeout_s=45)
            if not opened:
                opened = _open_multiupload_editor_by_index(
                    driver,
                    batch_start_idx + idx,
                    log_callback=log_callback,
                    timeout_s=45,
                )
            if not opened:
                # Fallback cuối: index tuyệt đối cũ (khi YouTube reorder bất thường).
                opened = _open_multiupload_editor_by_index(driver, idx, log_callback=log_callback, timeout_s=40)
            if not opened:
                out.append(
                    {
                        "success": False,
                        "url": None,
                        "error": f"Không bấm được Chỉnh sửa trong progress-list: {bn}",
                        "excel_done": False,
                    }
                )
                _log(
                    log_callback,
                    f"❌ Dừng lô: không mở được Chi tiết cho {bn}.",
                )
                _report_chrome_console_for_debug(
                    driver,
                    log_callback,
                    note=f"không mở Chỉnh sửa: {bn}",
                )
                for _p in paths_abs[idx + 1 :]:
                    out.append(
                        {
                            "success": False,
                            "url": None,
                            "error": "Bỏ qua sau lỗi trong cùng lô",
                            "excel_done": False,
                        }
                    )
                break

            _wait_upload_form_controls(driver, log_callback)
            _log(log_callback, "Đang chạy các bước trong Chi tiết (kids → next → bản quyền → visibility → lưu)...")

            def _on_link(url, _fp=path):
                if on_link_per_file:
                    try:
                        on_link_per_file(_fp, url)
                    except Exception:
                        pass

            r = _fill_metadata_to_done(
                driver,
                path,
                eff_title,
                made_for_kids,
                visibility,
                log_callback,
                _on_link,
            )
            out.append(r)
            if r.get("success"):
                _log(log_callback, f"✅ Đã hoàn tất Chi tiết cho {bn}.")
            else:
                _log(log_callback, f"❌ Chi tiết lỗi cho {bn}: {r.get('error') or 'unknown'}")
            _close_studio_editor_overlay(driver, log_callback)
            time.sleep(0.7)

            if not r.get("success"):
                _log(
                    log_callback,
                    f"❌ Dừng lô sau lỗi tại {bn}: {r.get('error') or 'unknown'}",
                )
                _report_chrome_console_for_debug(
                    driver,
                    log_callback,
                    note=f"Chi tiết/metadata lỗi: {bn}",
                )
                for _p in paths_abs[idx + 1 :]:
                    out.append(
                        {
                            "success": False,
                            "url": None,
                            "error": "Bỏ qua sau lỗi trong cùng lô",
                            "excel_done": False,
                        }
                    )
                break

    except Exception as e:
        _log(log_callback, f"Lỗi upload lô: {e}")
        _report_chrome_console_for_debug(driver, log_callback, note="exception upload_videos_batch")
        _dbg("B9", "upload_videos_batch exception", {"err": str(e)[:200]})
        while len(out) < len(paths_abs):
            out.append(
                {"success": False, "url": None, "error": str(e), "excel_done": False}
            )
    return out

def upload_video(
    driver,
    file_path,
    video_title="tool",
    made_for_kids=False,
    visibility="unlisted",
    log_callback=None,
    on_link_available=None,
):
    """
    Upload một video lên YouTube qua YouTube Studio.
    :param driver: WebDriver từ init_driver()
    :param file_path: Đường dẫn file video (mp4, mov, ...)
    :param video_title: Tiêu đề video
    :param made_for_kids: True = Made for kids
    :param visibility: 'public' | 'unlisted' | 'private'
    :param log_callback: Hàm callback(str) để ghi log
    :param on_link_available: Hàm callback(url) gọi khi có link (trước hoặc sau Done) để log + cập nhật Excel
    :return: dict {'success': bool, 'url': str|None, 'error': str|None, 'excel_done': bool}
    """
    if not driver or not os.path.isfile(file_path):
        return {"success": False, "url": None, "error": "Driver hoặc file không hợp lệ", "excel_done": False}

    file_path_abs = os.path.abspath(file_path)
    _dbg(
        "S0",
        "upload_video enter",
        {
            "file": os.path.basename(file_path_abs),
            "visibility": visibility,
            "hasTitle": bool(video_title and str(video_title).strip()),
            "madeForKids": bool(made_for_kids),
        },
    )
    try:
        file_input = _open_studio_and_get_file_input(driver, log_callback)
        file_input.send_keys(file_path_abs)
        _log(log_callback, "Đã chọn file video, đang chờ xử lý...")
        _wait_upload_form_controls(driver, log_callback)
        return _fill_metadata_to_done(
            driver,
            file_path_abs,
            video_title,
            made_for_kids,
            visibility,
            log_callback,
            on_link_available,
        )
    except Exception as e:
        _log(log_callback, f"Lỗi upload: {e}")
        _report_chrome_console_for_debug(driver, log_callback, note="exception upload_video")
        _dbg("S9", "upload_video exception", {"err": str(e)[:200]})
        return {"success": False, "url": None, "error": str(e), "excel_done": False}


def generate_excel(logs, excel_filename="YouTube_Upload_Links.xlsx", log_callback=None):
    """
    Tạo file Excel chứa danh sách link YouTube từ logs.
    Logs: list of dict {'timestamp': str, 'message': str}
    Tìm các dòng message chứa "Link: https://..."
    :return: Đường dẫn file Excel đã lưu, hoặc None.
    """
    if not OPENPYXL_AVAILABLE:
        _log(log_callback, "Thiếu thư viện openpyxl, không tạo được Excel.")
        return None

    # Parse links và tên file từ logs
    link_re = re.compile(r"Link:\s*(https?://[^\s]+)")
    upload_re = re.compile(r"Đang upload:\s*(.+)")
    success_re = re.compile(r"✅\s*Upload thành công:\s*(.+)")

    rows = []
    current_file = None
    for entry in logs:
        msg = entry.get("message", "")
        link_m = link_re.search(msg)
        upload_m = upload_re.search(msg)
        success_m = success_re.search(msg)
        if upload_m:
            current_file = upload_m.group(1).strip()
        if success_m:
            current_file = success_m.group(1).strip()
        if link_m:
            url = link_m.group(1).strip()
            if "youtube.com/watch" in url:
                url = url.split("&")[0]
            rows.append({"file": current_file or "N/A", "url": url})

    if not rows:
        _log(log_callback, "Không tìm thấy link nào trong logs để ghi Excel.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube Links"
    ws.append(["STT", "Tên file", "Link YouTube", "Thời gian", "Trạng thái"])
    for i, r in enumerate(rows, 1):
        ws.append([i, r["file"], r["url"], datetime.now().strftime("%Y-%m-%d %H:%M"), "SUCCESS"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max(len(str(c.value) or "") for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    output_dir = os.path.join(os.getcwd(), "output")
    os.makedirs(output_dir, exist_ok=True)
    safe_name = "".join(c for c in excel_filename if c.isalnum() or c in "._- ") or "YouTube_Upload_Links.xlsx"
    if not safe_name.endswith(".xlsx"):
        safe_name += ".xlsx"
    file_path = os.path.join(output_dir, safe_name)
    wb.save(file_path)
    return file_path
